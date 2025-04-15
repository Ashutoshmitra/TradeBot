from selenium import webdriver
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.common.exceptions import TimeoutException, NoSuchElementException, StaleElementReferenceException
from webdriver_manager.chrome import ChromeDriverManager
import time
import openpyxl
from datetime import datetime
import os
import re
import subprocess
import concurrent.futures
import threading
from functools import lru_cache

# Global variable for caching
_cached_options = {}

def setup_driver():
    """Set up the Chrome WebDriver with appropriate options."""
    options = webdriver.ChromeOptions()
    options.add_argument("--start-maximized")
    options.add_argument("--disable-extensions")
    options.add_argument("--disable-popup-blocking")
    options.add_argument("--disable-infobars")
    options.add_argument("--disable-dev-shm-usage")
    
    # Performance optimizations
    options.add_argument("--disable-notifications")
    options.add_argument("--disable-logging")
    options.add_argument("--log-level=3")
    options.add_argument("--disable-browser-side-navigation")
    options.add_argument("--disable-features=NetworkPrediction")
    options.add_argument("--disable-features=TranslateUI")
    options.add_argument("--disable-features=BlinkGenPropertyTrees")
    options.page_load_strategy = 'eager'  # Don't wait for all resources to load
    
    # Add headless mode options when running in GitHub Actions
    if os.environ.get('GITHUB_ACTIONS') == 'true':
        print("Running in GitHub Actions - enabling headless mode")
        options.add_argument('--headless')
        options.add_argument('--no-sandbox')
        options.add_argument('--disable-gpu')
        options.add_argument('--window-size=1920,1080')
    
    # Use webdriver-manager to handle ChromeDriver
    service = Service(ChromeDriverManager().install())
    driver = webdriver.Chrome(service=service, options=options)
    
    # Set script timeout
    driver.set_script_timeout(10)
    
    return driver

@lru_cache(maxsize=32)
def get_dropdown_options(driver, input_id):
    """Retrieve all available options from a dropdown with caching."""
    # Check if we have cached results
    cache_key = input_id
    if cache_key in _cached_options:
        return _cached_options[cache_key]
    
    try:
        dropdown = WebDriverWait(driver, 5).until(
            EC.element_to_be_clickable((By.ID, input_id))
        )
        driver.execute_script("arguments[0].scrollIntoView({block: 'center'});", dropdown)
        dropdown.click()
        time.sleep(0.5)  # Reduced sleep time
        
        # Use JavaScript for faster element selection
        options_js = """
            return Array.from(document.querySelectorAll('div[id^="{}"]'))
                .map(el => el.textContent.trim())
                .filter(text => text.length > 0);
        """.format(input_id.replace('input', 'option'))
        
        option_texts = driver.execute_script(options_js)
        
        # Close dropdown by clicking again
        dropdown.click()
        
        # Cache the results
        _cached_options[cache_key] = option_texts
        return option_texts
    except Exception as e:
        print(f"Error getting options for {input_id}: {e}")
        return []

def select_dropdown_option(driver, input_id, option_index, wait, trade_in_data=None):
    """Select an option from a dropdown by index using JavaScript for speed."""
    try:
        # Mostly JavaScript approach for speed
        option_selected = driver.execute_script(f"""
            var dropdown = document.getElementById('{input_id}');
            if (!dropdown) return false;
            
            dropdown.scrollIntoView({{block: 'center'}});
            dropdown.click();
            
            // Small delay to ensure dropdown opens
            await new Promise(resolve => setTimeout(resolve, 300));
            
            var optionId = '{input_id.replace("input", "option")}-{option_index}';
            var option = document.getElementById(optionId);
            
            if (!option) return false;
            
            var optionText = option.textContent.trim();
            option.click();
            
            return optionText;
        """)
        
        if not option_selected:
            # Fallback to the Selenium approach if JS fails
            dropdown = wait.until(EC.element_to_be_clickable((By.ID, input_id)))
            driver.execute_script("arguments[0].scrollIntoView({block: 'center'});", dropdown)
            dropdown.click()
            time.sleep(0.5)
            
            option_selector = f"//div[@id='{input_id.replace('input', 'option')}-{option_index}']"
            option = wait.until(EC.element_to_be_clickable((By.XPATH, option_selector)))
            option_text = option.text.strip()
            option.click()
            
            option_selected = option_text
        
        # Update trade-in data if provided
        if trade_in_data is not None and option_selected:
            if "react-select-2-input" in input_id:
                trade_in_data["Brand"] = option_selected
            elif "react-select-3-input" in input_id:
                trade_in_data["Model"] = option_selected
            elif "react-select-4-input" in input_id:
                trade_in_data["Variant"] = option_selected
                capacity_match = re.search(r'\d+\s*(?:GB|TB)(?=/|$)|(?<=/)\d+\s*(?:GB|TB)', option_selected)
                if capacity_match:
                    all_matches = re.findall(r'\d+\s*(?:GB|TB)', option_selected)
                    trade_in_data["Capacity"] = all_matches[-1] if all_matches else "Unknown"
                else:
                    trade_in_data["Capacity"] = "Unknown"
                trade_in_data["Type"] = "Smartphone"
        
        print(f"Selected option: {option_selected}")
        return True
    except Exception as e:
        print(f"Selection failed: {e}")
        return False

# Create a thread-safe version of the Excel saving function
_excel_lock = threading.Lock()

def save_to_excel(data):
    """Save the extracted trade-in data to an Excel file with locking for thread safety."""
    with _excel_lock:
        file_name = "tradein_values.xlsx"
        if os.path.exists(file_name):
            try:
                workbook = openpyxl.load_workbook(file_name)
                sheet = workbook.active
            except:
                workbook = openpyxl.Workbook()
                sheet = workbook.active
                headers = ["Country", "Device Type", "Brand", "Model", "Variant", "Capacity",
                           "Front Condition", "Value Type", "Currency", "Value", "Updated on"]
                sheet.append(headers)
        else:
            workbook = openpyxl.Workbook()
            sheet = workbook.active
            headers = ["Country", "Device Type", "Brand", "Model", "Variant", "Capacity",
                       "Front Condition", "Value Type", "Currency", "Value", "Updated on"]
            sheet.append(headers)

        row_data = [
            data["Country"],
            "Smartphone",
            data["Brand"],
            data["Model"],
            data.get("Variant", ""),
            data["Capacity"],
            data["Front Condition"],
            "Trade In",
            data["Currency"],
            data["Value"],
            data["Updated on"]
        ]
        sheet.append(row_data)

        try:
            workbook.save(file_name)
            print(f"Saved data to {file_name}")
        except Exception as e:
            print(f"Error saving Excel file: {e}")
            alt_file_name = f"tradein_values_{int(time.time())}.xlsx"
            try:
                workbook.save(alt_file_name)
                print(f"Saved to {alt_file_name}")
            except:
                print("Could not save Excel file")

def commit_changes(brand, model, variant, condition, value):
    """Commit changes to the repository after each row is added."""
    try:
        # Set up git config if running in GitHub Actions
        if os.environ.get('GITHUB_ACTIONS') == 'true':
            subprocess.run(["git", "config", "--global", "user.name", "GitHub Action Bot"])
            subprocess.run(["git", "config", "--global", "user.email", "action@github.com"])
        
        # Add the Excel file - using git operation queuing
        with _excel_lock:  # Use the same lock as Excel to prevent conflicts
            subprocess.run(["git", "add", "tradein_values.xlsx"])
            
            # Commit with a descriptive message
            commit_message = f"Add trade-in value for {brand} {model} {variant} ({condition}): {value}"
            result = subprocess.run(["git", "commit", "-m", commit_message], 
                                    capture_output=True, text=True)
            
            # Check if commit was successful
            if "nothing to commit" in result.stdout or "nothing to commit" in result.stderr:
                print("No changes to commit.")
                return False
            
            # Push the changes immediately
            subprocess.run(["git", "push"])
            print(f"Successfully committed and pushed changes: {commit_message}")
            return True
    
    except Exception as e:
        print(f"Error during git operations: {e}")
        return False

def navigate_and_complete_form(driver, wait, brand_index, model_index, variant_index, screen_condition):
    """Navigate the website and complete the form for a specific configuration."""
    screen_condition_id = f"LCDS-01-{screen_condition}"
    print(f"Processing: Brand index {brand_index}, Model index {model_index}, Variant index {variant_index}, Screen condition: {screen_condition}")

    # Initialize data dictionary
    trade_in_data = {
        "Country": "Singapore",
        "Device": "Smartphone",
        "Type": "Smartphone",
        "Brand": "",
        "Model": "",
        "Variant": "",
        "Capacity": "",
        "Front Condition": screen_condition.replace("_", " ").title(),
        "Value Type": "Trade In",
        "Currency": "",
        "Value": "",
        "Updated on": datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    }

    try:
        driver.get("https://compasiatradeinsg.com/tradein/sell")
        
        # Use JavaScript to click Smartphone button - faster than Selenium WebDriverWait
        smartphone_clicked = driver.execute_script("""
            var buttons = document.querySelectorAll('div.card-button-footer');
            for (var i = 0; i < buttons.length; i++) {
                if (buttons[i].textContent.trim() === 'Smartphone') {
                    buttons[i].click();
                    return true;
                }
            }
            return false;
        """)
        
        if not smartphone_clicked:
            # Fallback to Selenium
            wait.until(EC.element_to_be_clickable(
                (By.XPATH, "//div[contains(@class, 'card-button-footer') and text()='Smartphone']")
            )).click()
        
        time.sleep(0.5)  # Reduced sleep time

        # Select brand
        select_dropdown_option(driver, "react-select-2-input", brand_index, wait, trade_in_data)
        time.sleep(0.5)  # Reduced sleep time
        
        # Select model
        select_dropdown_option(driver, "react-select-3-input", model_index, wait, trade_in_data)
        time.sleep(0.5)  # Reduced sleep time
        
        # Select variant
        select_dropdown_option(driver, "react-select-4-input", variant_index, wait, trade_in_data)
        time.sleep(0.5)  # Reduced sleep time

        # Click next button using JavaScript for speed
        next_clicked = driver.execute_script("""
            var nextBtn = document.querySelector('button.progress-button-next:not([disabled])');
            if (nextBtn) {
                nextBtn.scrollIntoView({block: 'center'});
                nextBtn.click();
                return true;
            }
            return false;
        """)
        
        if not next_clicked:
            # Fallback to Selenium
            try:
                next_button = wait.until(EC.element_to_be_clickable(
                    (By.XPATH, "//button[contains(@class, 'progress-button-next') and not(@disabled)]")
                ))
                driver.execute_script("arguments[0].scrollIntoView({block: 'center'});", next_button)
                next_button.click()
            except Exception as e:
                print(f"Failed to click Next button: {e}")
                return False

        time.sleep(1)  # Reduced sleep time

        # Fill diagnostic form using JavaScript - using a more efficient combined JavaScript execution
        form_filled = driver.execute_script(f"""
            function clickYesButton(labelText) {{
                var labels = document.querySelectorAll('label.diagnostic-form-label');
                for (var i = 0; i < labels.length; i++) {{
                    if (labels[i].textContent.includes(labelText)) {{
                        var yesButton = labels[i].parentNode.querySelector('button:first-of-type');
                        if (yesButton) {{
                            yesButton.click();
                            return true;
                        }}
                        return false;
                    }}
                }}
                return false;
            }}
            
            // Track overall success
            let success = true;
            
            // Device locks
            success = success && clickYesButton('Is your device free of any locks');
            
            // Screen condition
            var screenLabel = document.querySelector('label[for="{screen_condition_id}"]');
            if (screenLabel) {{
                screenLabel.click();
            }} else {{
                success = false;
            }}
            
            // Body condition
            var bodyLabel = document.querySelector('label[for="DECO-01-flawless"]');
            if (bodyLabel) {{
                bodyLabel.click();
            }} else {{
                success = false;
            }}
            
            // Other conditions
            success = success && clickYesButton('Fingerprint/Face ID working');
            success = success && clickYesButton('device functions below working fine');
            success = success && clickYesButton('front and back cameras');
            
            // None of the above checkbox
            let foundCheckbox = false;
            var labels = document.querySelectorAll('label');
            for (var i = 0; i < labels.length; i++) {{
                if (labels[i].textContent.trim().includes('None of the above')) {{
                    var checkbox = labels[i].previousElementSibling;
                    if (!checkbox || checkbox.type !== 'checkbox') {{
                        var parent = labels[i].parentElement;
                        checkbox = parent.querySelector('input[type="checkbox"]');
                    }}
                    if (checkbox && checkbox.type === 'checkbox') {{
                        checkbox.checked = true;
                        checkbox.click();
                        checkbox.dispatchEvent(new Event('change', {{ bubbles: true }}));
                        foundCheckbox = true;
                        break;
                    }}
                }}
            }}
            
            success = success && foundCheckbox;
            
            // Click Get Quote button after a small delay to ensure form is processed
            setTimeout(() => {{
                var quoteBtn = document.querySelector('button[type="submit"]');
                if (quoteBtn) {{
                    quoteBtn.scrollIntoView({{block: 'center'}});
                    quoteBtn.click();
                }}
            }}, 500);
            
            return success;
        """)
        
        if not form_filled:
            print("Warning: Form may not be completely filled. Continuing anyway...")
            
            # Try to click Get Quote button if it wasn't clicked by the JavaScript
            try:
                quote_button = wait.until(EC.element_to_be_clickable(
                    (By.XPATH, "//button[@type='submit' and contains(text(), 'Get Quote')]")
                ))
                driver.execute_script("arguments[0].scrollIntoView({block: 'center'});", quote_button)
                quote_button.click()
            except Exception as e:
                print(f"Failed to click Get Quote button: {e}")
                driver.execute_script("""
                    var quoteBtn = document.querySelector('button[type="submit"]');
                    if (quoteBtn) {
                        quoteBtn.scrollIntoView({block: 'center'});
                        quoteBtn.click();
                    }
                """)

        # Wait for results with a more aggressive timeout
        wait_for_results = WebDriverWait(driver, 10, 0.5)
        try:
            wait_for_results.until(EC.visibility_of_element_located((By.CLASS_NAME, "pricing-display-table")))
            
            # Extract trade-in value using JavaScript for speed
            trade_in_result = driver.execute_script("""
                var result = { currency: "", price: "" };
                
                var currencyElement = document.querySelector(".price-product-name.currency");
                if (currencyElement) {
                    result.currency = currencyElement.textContent.trim();
                } else {
                    result.currency = "SGD";  // Default if not found
                }
                
                var priceElement = document.querySelector(".pricing-display-price");
                if (priceElement) {
                    var priceText = priceElement.textContent.trim();
                    result.price = priceText.replace(/[^0-9.]/g, '');
                }
                
                return result;
            """)
            
            if trade_in_result and trade_in_result.get("price"):
                trade_in_data["Currency"] = trade_in_result.get("currency", "SGD")
                trade_in_data["Value"] = trade_in_result.get("price", "")
                
                print(f"Extracted trade-in value: {trade_in_data['Currency']} {trade_in_data['Value']}")
                
                # Save data in a non-blocking way
                save_to_excel(trade_in_data)
                
                # Commit changes to the repository after each successful scrape
                commit_changes(
                    trade_in_data["Brand"], 
                    trade_in_data["Model"], 
                    trade_in_data["Variant"], 
                    trade_in_data["Front Condition"], 
                    f"{trade_in_data['Currency']} {trade_in_data['Value']}"
                )
                
                return True
            else:
                print("Failed to extract trade-in value: No price found")
                return False

        except Exception as e:
            print(f"Failed to extract trade-in value: {e}")
            if os.environ.get('GITHUB_ACTIONS') == 'true':
                driver.save_screenshot(f"error_{brand_index}_{model_index}_{variant_index}_{screen_condition}.png")
            return False

    except Exception as main_error:
        print(f"Main execution failed: {main_error}")
        if os.environ.get('GITHUB_ACTIONS') == 'true':
            driver.save_screenshot(f"error_{brand_index}_{model_index}_{variant_index}_{screen_condition}.png")
        return False

def process_configuration(brand_idx, model_idx, variant_idx, condition):
    """Process a single configuration using its own browser instance."""
    driver = None
    try:
        driver = setup_driver()
        ignored_exceptions = (NoSuchElementException, StaleElementReferenceException)
        wait = WebDriverWait(driver, 10, 0.5, ignored_exceptions=ignored_exceptions)
        
        result = navigate_and_complete_form(driver, wait, brand_idx, model_idx, variant_idx, condition)
        return result
    except Exception as e:
        print(f"Error in configuration {brand_idx}_{model_idx}_{variant_idx}_{condition}: {e}")
        return False
    finally:
        if driver:
            driver.quit()

def test_with_first_five():
    """Test with just the first 5 configurations using parallel processing."""
    # Setup a driver just to get the available options
    driver = setup_driver()
    ignored_exceptions = (NoSuchElementException, StaleElementReferenceException)
    wait = WebDriverWait(driver, 15, 0.5, ignored_exceptions=ignored_exceptions)

    try:
        # Take screenshot of initial state for debugging
        if os.environ.get('GITHUB_ACTIONS') == 'true':
            driver.save_screenshot("initial_page.png")
            
        # Navigate to website and click on Smartphone
        driver.get("https://compasiatradeinsg.com/tradein/sell")
        wait.until(EC.element_to_be_clickable(
            (By.XPATH, "//div[contains(@class, 'card-button-footer') and text()='Smartphone']")
        )).click()
        time.sleep(1)

        # We'll just test with Apple brand (usually index 0)
        brand_idx = 0
        
        # Select the brand
        select_dropdown_option(driver, "react-select-2-input", brand_idx, wait)
        time.sleep(1)
        
        # Get available models for this brand
        model_options = get_dropdown_options(driver, "react-select-3-input")
        print(f"Available models: {model_options}")
        
        # Just test with first model
        model_idx = 0
        
        # For the model, go back to the main page
        driver.get("https://compasiatradeinsg.com/tradein/sell")
        wait.until(EC.element_to_be_clickable(
            (By.XPATH, "//div[contains(@class, 'card-button-footer') and text()='Smartphone']")
        )).click()
        time.sleep(1)
        
        # Select brand again
        select_dropdown_option(driver, "react-select-2-input", brand_idx, wait)
        time.sleep(1)
        
        # Select model
        select_dropdown_option(driver, "react-select-3-input", model_idx, wait)
        time.sleep(1)
        
        # Get variants for this model
        variant_options = get_dropdown_options(driver, "react-select-4-input")
        print(f"Available variants: {variant_options}")
        
        # For testing, we'll just do 5 combinations
        test_combinations = [
            (0, "flawless"),        # First variant, flawless screen
            (0, "minor_scratches"), # First variant, minor scratches
            (1, "flawless"),        # Second variant, flawless screen
            (1, "minor_scratches"), # Second variant, minor scratches
            (0, "cracked")          # First variant, cracked screen
        ]
        
    except Exception as e:
        print(f"Error gathering options: {e}")
        test_combinations = [
            (0, "flawless"),
            (0, "minor_scratches"),
            (1, "flawless"),
            (1, "minor_scratches"),
            (0, "cracked")
        ]
    finally:
        driver.quit()
    
    # Process the configurations in parallel
    print(f"Processing {len(test_combinations)} configurations in parallel")
    
    # Detect number of available CPUs and set appropriate max_workers
    max_workers = min(os.cpu_count() or 4, 3)  # Use at most 3 workers to avoid rate limiting
    
    with concurrent.futures.ThreadPoolExecutor(max_workers=max_workers) as executor:
        futures = []
        for variant_idx, condition in test_combinations:
            print(f"Submitting: Brand idx {brand_idx}, Model idx {model_idx}, Variant idx {variant_idx}, Condition: {condition}")
            future = executor.submit(process_configuration, brand_idx, model_idx, variant_idx, condition)
            futures.append(future)
        
        # Wait for results
        completed = 0
        for future in concurrent.futures.as_completed(futures):
            result = future.result()
            if result:
                completed += 1
                print(f"Successfully processed configuration {completed}")
            else:
                print(f"Failed to process configuration")
    
    print(f"Test run complete. Successfully processed {completed}/{len(test_combinations)} configurations.")

if __name__ == "__main__":
    test_with_first_five()