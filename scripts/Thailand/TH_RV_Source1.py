from selenium import webdriver
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.common.exceptions import TimeoutException, NoSuchElementException, StaleElementReferenceException
from webdriver_manager.chrome import ChromeDriverManager
import time
import openpyxl
from datetime import datetime
import os
import re
import argparse

def setup_driver(headless=True):
    """Set up the Chrome WebDriver with appropriate options for containerized environment."""
    options = webdriver.ChromeOptions()
    
    # Only enable headless mode if specified
    if headless:
        options.add_argument('--headless')
    
    # Core options for stability
    options.add_argument('--no-sandbox')
    options.add_argument('--disable-dev-shm-usage')
    options.add_argument('--disable-gpu')
    options.add_argument('--start-maximized')
    options.add_argument('--window-size=1920,1080')
    
    # Performance options
    options.add_argument('--disable-extensions')
    options.add_argument('--disable-infobars')
    options.add_argument('--disable-logging')
    options.add_argument('--disable-notifications')
    options.add_argument('--enable-javascript')
    
    # Cache settings
    options.add_argument('--disk-cache-size=1048576')
    options.add_argument('--media-cache-size=1048576')
    
    # User agent
    options.add_argument('--user-agent=Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/91.0.4472.124 Safari/537.36')
    
    # Use a normal page load strategy instead of eager
    options.page_load_strategy = 'normal'
    
    # Initialize the driver directly
    driver = webdriver.Chrome(options=options)
    
    # Set page load timeout to be more generous
    driver.set_page_load_timeout(60)
    
    return driver


def handle_city_modal(driver, wait):
    """Handle the city selection modal that appears on page load."""
    try:
        # Wait for the modal to appear
        print("Checking for city selection modal...")
        time.sleep(2)
        
        # Check if the modal is present by looking for the select element
        try:
            # First try the desktop version
            select_element = driver.find_element(By.CSS_SELECTOR, "div.ant-select-selection")
            print("Found desktop city selector")
            
            # Click on the select element to open dropdown
            select_element.click()
            time.sleep(1)
            
            # Find the first option (should be Bangkok)
            first_option = wait.until(
                EC.element_to_be_clickable((By.CSS_SELECTOR, "li.ant-select-dropdown-menu-item:first-child"))
            )
            first_option.click()
            time.sleep(1)
            
        except:
            # Try the mobile version
            try:
                print("Trying mobile city selector...")
                mobile_select = driver.find_element(By.CSS_SELECTOR, "select.SelectCommon__StyledSelectCommon-sc-17ycdhb-0")
                
                # Use JavaScript to select the first option
                driver.execute_script("""
                    var select = arguments[0];
                    select.selectedIndex = 1; // Select the first option after the placeholder
                    var event = new Event('change', { bubbles: true });
                    select.dispatchEvent(event);
                """, mobile_select)
                time.sleep(1)
                
            except Exception as e:
                print(f"Could not find city selector: {e}")
        
        # Click the "เริ่มเลย" (Start) button
        try:
            start_button = wait.until(
                EC.element_to_be_clickable((By.CSS_SELECTOR, "button.ant-btn.ant-btn-primary.ant-btn-lg"))
            )
            start_button.click()
            time.sleep(2)
            print("Successfully clicked start button and closed city modal")
            return True
            
        except Exception as e:
            print(f"Could not find start button: {e}")
            return False
            
    except Exception as e:
        print(f"No city modal found or error handling it: {e}")
        return True  # Continue anyway as the modal might not always appear


def get_condition_mapping(screen_condition):
    """Map screen condition values to the required conditions."""
    condition_map = {
        "No scratches": "Flawless",
        "There are visible marks but they are not clear.": "Good",
        "There are visible scratches.": "Damaged"
    }
    return condition_map.get(screen_condition, screen_condition)


def detect_device_type(device_name):
    """Determine if the device is a tablet or smartphone based on the name."""
    device_name_lower = device_name.lower()
    
    # Check for tablet keywords
    if any(keyword in device_name_lower for keyword in ["ipad", "tab", "tablet", "galaxy tab"]):
        return "Tablet"
    elif any(keyword in device_name_lower for keyword in ["watch"]):
        return "SmartWatch"
    else:
        # Default to Smartphone for other devices
        return "SmartPhone"


def click_brand_tab(driver, wait, brand_name):
    """Click on a specific brand tab (Apple or Samsung)."""
    # For Apple, we don't need to click anything as it's displayed by default
    if brand_name.lower() == "apple":
        print(f"Apple phones are displayed by default")
        time.sleep(2)  # Wait for models to load
        return True
    
    # For Samsung, we need to click the Samsung tab
    if brand_name.lower() == "samsung":
        try:
            # Find all brand tabs - they are in the home__StyledSection-sc-d27gbn-0 section
            # Look for the Samsung tab specifically
            samsung_tab = wait.until(
                EC.element_to_be_clickable((By.XPATH, "//div[contains(@class, 'home__StyledModelSelectionTab-sc-d27gbn-4')]//img[contains(@src, 'samsung_logo')]//parent::div//parent::div"))
            )
            
            # Scroll to the element and click
            driver.execute_script("arguments[0].scrollIntoView({block: 'center'});", samsung_tab)
            time.sleep(0.5)
            samsung_tab.click()
            time.sleep(3)  # Wait for the Samsung models to load
            print(f"Successfully clicked on Samsung tab")
            return True
            
        except Exception as e:
            print(f"Error clicking Samsung tab: {e}")
            # Try alternative method
            try:
                # Look for the SAMSUNG text in the tabs
                samsung_tab = driver.find_element(By.XPATH, "//div[contains(@class, 'home__StyledModelSelectionTab-sc-d27gbn-4') and contains(., 'SAMSUNG')]")
                driver.execute_script("arguments[0].scrollIntoView({block: 'center'});", samsung_tab)
                samsung_tab.click()
                time.sleep(3)
                print(f"Successfully clicked on Samsung tab (alternative method)")
                return True
            except Exception as e2:
                print(f"Both methods failed to click Samsung tab: {e2}")
                return False
    
    return False


def get_smartphone_cards(driver, wait):
    """Get all smartphone cards for the currently selected brand."""
    try:
        # Wait for the smartphone cards to load - using the specific class from the HTML
        smartphone_cards = wait.until(
            EC.presence_of_all_elements_located((By.CSS_SELECTOR, "div[class*='ProductCard__StyledProductCard-sc-1msc4z7-0'][href]"))
        )
        
        smartphone_info = []
        for card in smartphone_cards:
            try:
                # Get the title
                title_element = card.find_element(By.CSS_SELECTOR, "div.title")
                title = title_element.text.strip()
                
                # Get the href attribute (the div has href attribute)
                href = card.get_attribute("href")
                
                # Clean up the href to create full URL
                if href and href.startswith("/"):
                    # Replace spaces with underscores as shown in your example
                    href_cleaned = href.replace(" ", "_")
                    full_url = f"https://www.remobie.com{href_cleaned}"
                else:
                    continue  # Skip if no valid href
                
                # Determine device type
                device_type = detect_device_type(title)
                
                smartphone_info.append({
                    "title": title,
                    "url": full_url,
                    "device_type": device_type
                })
                
            except Exception as e:
                print(f"Error extracting card info: {e}")
                continue
        
        print(f"Found {len(smartphone_info)} device cards")
        return smartphone_info
        
    except Exception as e:
        print(f"Error getting device cards: {e}")
        return []


def click_form_option(driver, wait, option_text, question_number):
    """Click on a specific option in the form question."""
    try:
        print(f"DEBUG: Looking for option '{option_text}' in question {question_number}")
        
        # Wait for the question section to be present
        time.sleep(0.5)  # Reduced wait
        
        # Find all condition items in the current question
        condition_items = driver.find_elements(By.CSS_SELECTOR, "div.condition-item")
        
        for item in condition_items:
            try:
                title_element = item.find_element(By.CSS_SELECTOR, "div.title")
                if title_element.text.strip() == option_text:
                    print(f"DEBUG: Found option '{option_text}', clicking...")
                    driver.execute_script("arguments[0].scrollIntoView({block: 'center'});", title_element)
                    driver.execute_script("arguments[0].click();", title_element)
                    
                    # Wait for a short time to let the next question appear
                    time.sleep(0.3)
                    print(f"DEBUG: Successfully clicked '{option_text}'")
                    return True
            except Exception as e:
                continue
        
        print(f"DEBUG: Could not find option '{option_text}' in question {question_number}")
        return False
        
    except Exception as e:
        print(f"DEBUG: Error clicking option '{option_text}': {e}")
        return False


def get_storage_options(driver, wait):
    """Get all available storage options for the phone."""
    try:
        print("DEBUG: Getting storage options...")
        storage_options = []
        
        # Find all condition items that likely contain storage options
        condition_items = driver.find_elements(By.CSS_SELECTOR, "div.condition-item")
        
        for item in condition_items:
            try:
                title_element = item.find_element(By.CSS_SELECTOR, "div.title")
                text = title_element.text.strip()
                # Check if it's a storage option (contains GB)
                if "GB" in text:
                    storage_options.append(text)
            except:
                continue
        
        print(f"DEBUG: Found storage options: {storage_options}")
        return storage_options
        
    except Exception as e:
        print(f"DEBUG: Error getting storage options: {e}")
        return []


def click_view_price_button(driver, wait):
    """Click the 'View price offers' button."""
    try:
        print("DEBUG: Looking for 'View price offers' button...")
        
        # First try to find the Thai version
        button_selectors = [
            "//button[contains(@class, 'ant-btn-primary')]//span[contains(text(), 'ดูข้อเสนอราคา')]",
            "//button[contains(@class, 'ant-btn-primary')]//span[contains(text(), 'View price offers')]",
            "//button[contains(@class, 'ant-btn-primary') and contains(text(), 'ดูข้อเสนอราคา')]",
            "//button[contains(@class, 'ant-btn-primary') and contains(text(), 'View price offers')]"
        ]
        
        for selector in button_selectors:
            try:
                button = driver.find_element(By.XPATH, selector)
                if button:
                    print(f"DEBUG: Found button with selector: {selector}")
                    driver.execute_script("arguments[0].scrollIntoView({block: 'center'});", button)
                    driver.execute_script("arguments[0].click();", button)
                    time.sleep(3)
                    print("DEBUG: Successfully clicked 'View price offers' button")
                    return True
            except:
                continue
        
        # If not found by text, try by class and position
        try:
            buttons = driver.find_elements(By.CSS_SELECTOR, "button.ant-btn.ant-btn-primary.ant-btn-lg")
            if len(buttons) >= 2:
                # The second button should be the "View price offers" button
                button = buttons[1]
                driver.execute_script("arguments[0].scrollIntoView({block: 'center'});", button)
                driver.execute_script("arguments[0].click();", button)
                time.sleep(3)
                print("DEBUG: Successfully clicked 'View price offers' button (by position)")
                return True
        except Exception as e:
            print(f"DEBUG: Error with position-based selection: {e}")
        
        print("DEBUG: Could not find 'View price offers' button")
        return False
        
    except Exception as e:
        print(f"DEBUG: Error clicking 'View price offers' button: {e}")
        return False


def extract_trade_in_price(driver, wait):
    """Extract the trade-in price from the price page."""
    try:
        print("DEBUG: Extracting trade-in price...")
        
        # Wait for the price section to load
        price_section = wait.until(
            EC.presence_of_element_located((By.CSS_SELECTOR, "div.offer__StyledOfferPrice-sc-2lrby1-0"))
        )
        
        # Extract the price
        price_element = price_section.find_element(By.TAG_NAME, "h1")
        price_text = price_element.text.strip()
        
        # Clean up the price (remove non-numeric characters except decimal point)
        price_clean = re.sub(r'[^\d.]', '', price_text.split()[0])
        
        print(f"DEBUG: Extracted price: {price_clean} baht")
        return price_clean
        
    except Exception as e:
        print(f"DEBUG: Error extracting trade-in price: {e}")
        return None


def fill_trade_in_form(driver, wait, smartphone_data, storage, screen_condition, max_retries=3):
    """Fill the trade-in form for a specific configuration with retry logic."""
    for attempt in range(max_retries):
        try:
            print(f"DEBUG: Attempt {attempt + 1}/{max_retries} for {storage}/{screen_condition}")
            
            # Question 1: Select storage capacity
            print(f"DEBUG: Filling Q1 - Storage: {storage}")
            if not click_form_option(driver, wait, storage, 1):
                continue
            
            # Question 2: Can the phone be turned on? Select "can"
            print("DEBUG: Filling Q2 - Can phone be turned on: ได้")
            if not click_form_option(driver, wait, "ได้", 2):
                if not click_form_option(driver, wait, "can", 2):  # Try English version
                    continue
            
            # Question 3: Can I sign out of iCloud/Google? Select "can" 
            print("DEBUG: Filling Q3 - Can sign out of iCloud/Google: ได้")
            if not click_form_option(driver, wait, "ได้", 3):
                if not click_form_option(driver, wait, "can", 3):  # Try English version
                    continue
            
            # Question 4: Screen condition - Map to Thai options
            thai_screen_conditions = {
                "No scratches": "ไม่มีรอยขีดข่วน",
                "There are visible marks but they are not clear.": "มีรอยเห็นได้แต่ไม่ชัด",
                "There are visible scratches.": "ไม่มีรอยขีดข่วน"  # Changed to use "No scratches" for damaged
            }

            thai_condition = thai_screen_conditions.get(screen_condition, screen_condition)
            print(f"DEBUG: Filling Q4 - Screen condition: {thai_condition}")
            if not click_form_option(driver, wait, thai_condition, 4):
                # For damaged condition with English fallback, use "No scratches"
                if screen_condition == "There are visible scratches.":
                    if not click_form_option(driver, wait, "No scratches", 4):
                        continue
                # Try original English version as fallback for others
                elif not click_form_option(driver, wait, screen_condition, 4):
                    continue
            
            # Question 5: Country - Select "Thai"
            print("DEBUG: Filling Q5 - Country: ไทย")
            if not click_form_option(driver, wait, "ไทย", 5):
                if not click_form_option(driver, wait, "Thai", 5):  # Try English version
                    continue
            
            # Question 6: Phone problems - Select "do not have" or "Broken Screen" based on condition
            if screen_condition == "There are visible scratches." or get_condition_mapping(screen_condition) == "Damaged":
                print("DEBUG: Filling Q6 - Phone problems: หน้าจอแตก")
                if not click_form_option(driver, wait, "หน้าจอแตก", 6):
                    if not click_form_option(driver, wait, "Broken Screen", 6):  # Try English version
                        continue
            else:
                print("DEBUG: Filling Q6 - Phone problems: ไม่มี")
                if not click_form_option(driver, wait, "ไม่มี", 6):
                    if not click_form_option(driver, wait, "do not have", 6):  # Try English version
                        continue
            
            # Question 7: Leave blank (no selection needed)
            print("DEBUG: Skipping Q7 (no selection needed)")
            
            # Wait a bit before clicking the button
            time.sleep(0.5)
            
            # Click "View price offers" button
            if not click_view_price_button(driver, wait):
                continue
                
            # Extract the price
            price = extract_trade_in_price(driver, wait)
            if price:
                smartphone_data["Value"] = price
                smartphone_data["Condition"] = get_condition_mapping(screen_condition)
                smartphone_data["Capacity"] = storage
                
                return True
            
        except Exception as e:
            print(f"DEBUG: Error in fill_trade_in_form (attempt {attempt + 1}/{max_retries}): {e}")
            if attempt == max_retries - 1:  # Last attempt
                print(f"DEBUG: Final attempt failed, giving up on this configuration")
                # Set condition value even if price extraction failed
                smartphone_data["Condition"] = get_condition_mapping(screen_condition)
                smartphone_data["Capacity"] = storage
                smartphone_data["Value"] = "NA"
                return False
            else:
                print(f"DEBUG: Retrying... (attempt {attempt + 2}/{max_retries})")
                time.sleep(2)
                continue
    
    return False


def save_to_excel(data, output_file):
    """Save the extracted trade-in data to an Excel file."""
    # Ensure output directory exists
    os.makedirs(os.path.dirname(output_file) if os.path.dirname(output_file) else ".", exist_ok=True)
    
    if os.path.exists(output_file):
        try:
            workbook = openpyxl.load_workbook(output_file)
            sheet = workbook.active
        except:
            workbook = openpyxl.Workbook()
            sheet = workbook.active
            # Use the new column headers
            headers = ["Country", "Device Type", "Brand", "Model", "Capacity", "Color", 
                      "Launch RRP", "Condition", "Value Type", "Currency", "Value", 
                      "Source", "Updated on", "Updated by", "Comments"]
            sheet.append(headers)
    else:
        workbook = openpyxl.Workbook()
        sheet = workbook.active
        # Use the new column headers
        headers = ["Country", "Device Type", "Brand", "Model", "Capacity", "Color", 
                  "Launch RRP", "Condition", "Value Type", "Currency", "Value", 
                  "Source", "Updated on", "Updated by", "Comments"]
        sheet.append(headers)

    # Add new row with all the required columns
    row_data = [
        data["Country"],
        data["Device Type"],
        data["Brand"],
        data["Model"],
        data["Capacity"],
        data["Color"],
        data["Launch RRP"],
        data["Condition"],
        data["Value Type"],
        data["Currency"],
        data["Value"],
        data["Source"],
        data["Updated on"],
        data["Updated by"],
        data["Comments"]
    ]
    sheet.append(row_data)

    try:
        workbook.save(output_file)
        print(f"Saved data to {output_file}")
    except Exception as e:
        print(f"Error saving Excel file: {e}")
        alt_file_name = os.path.join(os.path.dirname(output_file), f"TH_RV_Source1.xlsx")
        try:
            workbook.save(alt_file_name)
            print(f"Saved to {alt_file_name}")
        except:
            print("Could not save Excel file")


def process_smartphone_configuration(driver, wait, smartphone, storage, screen_condition, output_file):
    """Process a single smartphone configuration with retry logic."""
    max_retries = 3
    
    for attempt in range(max_retries):
        try:
            print(f"\n--- Storage: {storage}, Condition: {screen_condition} (Attempt {attempt + 1}/{max_retries}) ---")
            
            # Create data dictionary for this configuration
            trade_in_data = {
                "Country": "Thailand",
                "Device Type": smartphone.get("device_type", "SmartPhone"),
                "Brand": smartphone.get("brand", ""),
                "Model": smartphone['title'],
                "Capacity": storage,
                "Color": "",
                "Launch RRP": "",
                "Condition": "",
                "Value Type": "Trade-in",
                "Currency": "THB",
                "Value": "",
                "Source": "TH_RV_Source1",
                "Updated on": datetime.now().strftime("%Y-%m-%d"),
                "Updated by": "",
                "Comments": ""
            }
            
            # Navigate back to the smartphone page for fresh form
            driver.get(smartphone['url'])
            time.sleep(1.5)
            
            # Fill the form and get the price
            if fill_trade_in_form(driver, wait, trade_in_data, storage, screen_condition):
                print(f"Successfully extracted trade-in value: {trade_in_data['Value']} THB")
                save_to_excel(trade_in_data, output_file)
                return True
            else:
                print(f"Failed to extract trade-in value for {storage} / {screen_condition}")
                if attempt == max_retries - 1:  # Last attempt
                    # Make sure condition is set even if price extraction failed
                    trade_in_data["Condition"] = get_condition_mapping(screen_condition)
                    trade_in_data["Value"] = "NA"
                    save_to_excel(trade_in_data, output_file)
                    return False
        
        except Exception as e:
            print(f"ERROR: {e}")
            if attempt == max_retries - 1:  # Last attempt
                print(f"Final attempt failed for {storage}/{screen_condition}, skipping...")
                # Save error record with condition set and NA for value
                trade_in_data = {
                    "Country": "Thailand",
                    "Device Type": smartphone.get("device_type", "SmartPhone"),
                    "Brand": smartphone.get("brand", ""),
                    "Model": smartphone['title'],
                    "Capacity": storage,
                    "Color": "",
                    "Launch RRP": "",
                    "Condition": get_condition_mapping(screen_condition),
                    "Value Type": "Trade-in",
                    "Currency": "THB",
                    "Value": "NA",
                    "Source": "TH_RV_Source1",
                    "Updated on": datetime.now().strftime("%Y-%m-%d"),
                    "Updated by": "",
                    "Comments": ""
                }
                save_to_excel(trade_in_data, output_file)
                return False
            else:
                print(f"Retrying... (attempt {attempt + 2}/{max_retries})")
                time.sleep(3)
                continue
    
    return False


def main_loop(n_scrape=None, output_file=None):
    """Main loop to iterate through brands and smartphone models."""
    brands = ["Apple", "Samsung"]
    screen_conditions = [
        "No scratches",
        "There are visible marks but they are not clear.",
        "There are visible scratches."
    ]
    
    # Use default output path if not specified
    if output_file is None:
        # Check for environment variable first
        output_dir = os.environ.get("OUTPUT_DIR", "output")
        os.makedirs(output_dir, exist_ok=True)
        output_file = os.path.join(output_dir, "TH_RV_Source1.xlsx")
    
    print(f"Will save results to: {output_file}")

    # Setup driver (single browser instance)
    driver = setup_driver(headless=True)
    
    ignored_exceptions = (NoSuchElementException, StaleElementReferenceException)
    wait = WebDriverWait(driver, 15, 0.5, ignored_exceptions=ignored_exceptions)

    # Counter for tracking scrapes
    total_scrape_count = 0
    
    try:
        for brand in brands:
            print(f"\n========== Processing {brand} ==========\n")
            
            # Navigate to the main page
            driver.get("https://www.remobie.com/")
            time.sleep(2)
            
            # Handle city selection modal only when loading the homepage
            handle_city_modal(driver, wait)
            
            # Click on the brand tab (or use default for Apple)
            if not click_brand_tab(driver, wait, brand):
                print(f"Could not click on {brand} tab, skipping")
                continue
            
            # Get all smartphone cards for this brand
            smartphone_cards = get_smartphone_cards(driver, wait)
            print(f"Found {len(smartphone_cards)} devices for {brand}")
            
            # Process each smartphone
            for smartphone in smartphone_cards:
                print(f"\n=== Processing: {smartphone['title']} ===")
                print(f"URL: {smartphone['url']}")
                print(f"Device Type: {smartphone.get('device_type', 'SmartPhone')}")
                
                # Add brand to smartphone data
                smartphone["brand"] = brand
                
                # Navigate to the smartphone page
                driver.get(smartphone['url'])
                time.sleep(2)
                
                # Get storage options for this smartphone
                storage_options = get_storage_options(driver, wait)
                if not storage_options:
                    print(f"DEBUG: No storage options found for {smartphone['title']}, skipping...")
                    continue
                
                # Process each storage option
                for storage in storage_options:
                    print(f"\n--- Processing storage: {storage} ---")
                    
                    # Process each screen condition
                    for screen_condition in screen_conditions:
                        # Process the configuration with retry logic
                        success = process_smartphone_configuration(
                            driver, wait, smartphone, storage, screen_condition, output_file
                        )
                        
                        total_scrape_count += 1
                        
                        # Check if we've reached the requested number of scrapes
                        if n_scrape is not None and total_scrape_count >= n_scrape:
                            print(f"Completed {n_scrape} scrapes as requested.")
                            return
                        
                        # Brief pause between configurations
                        time.sleep(1)

    except Exception as e:
        print(f"Error in main loop: {e}")
        import traceback
        traceback.print_exc()
        
        error_folder = os.path.dirname(output_file)
        os.makedirs(error_folder, exist_ok=True)
        driver.save_screenshot(os.path.join(error_folder, "main_loop_error.png"))
    finally:
        driver.quit()
        print("Browser closed. Process complete.")


if __name__ == "__main__":
    parser = argparse.ArgumentParser(description='Scrape trade-in values for devices from Thailand Remobie website')
    parser.add_argument('-n', type=int, help='Number of scrapes to perform (e.g., -n 2 will scrape 2 devices)', default=None)
    parser.add_argument('-o', '--output', type=str, help='Output Excel file path', default=None)
    args = parser.parse_args()
    
    main_loop(args.n, args.output)