# This script should be run in head mode strictly with a GUI
from selenium import webdriver
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait, Select
from selenium.webdriver.support import expected_conditions as EC
from selenium.common.exceptions import TimeoutException, NoSuchElementException, ElementClickInterceptedException
import time
import logging
import os
import re
import openpyxl  # Added import for Excel functionality
from datetime import datetime
import argparse

# Set up logging
def setup_logging(log_file=None):
    """Set up logging to console only."""
    logger = logging.getLogger()
    logger.setLevel(logging.INFO)
    
    # Remove existing handlers if any (prevent duplicate logs)
    if logger.hasHandlers():
        logger.handlers.clear()
    
    # Create formatter
    formatter = logging.Formatter('%(asctime)s - %(levelname)s - %(message)s')
    
    # Create console handler
    console_handler = logging.StreamHandler()
    console_handler.setFormatter(formatter)
    
    # Add only the console handler
    logger.addHandler(console_handler)
    
    logging.info("Logging to console only")
    return None


def setup_driver(headless=False):
    """Set up the Chrome WebDriver with minimal options."""
    options = webdriver.ChromeOptions()
    
    # Core options
    if headless:
        options.add_argument('--headless=new')
        # Add these options for better headless performance
        options.add_argument('--disable-features=VizDisplayCompositor')
        options.add_argument('--disable-gpu-sandbox')
        options.add_argument('--disable-software-rasterizer')
        options.add_argument('--window-size=1920,1080')  # Ensure a good viewport size
        options.add_argument('--force-device-scale-factor=1')
        # Force enable JavaScript
        options.add_argument('--enable-javascript')
        # Disable web security for better rendering
        options.add_argument('--disable-web-security')
    
    options.add_argument('--disable-gpu')
    options.add_argument('--no-sandbox')
    options.add_argument('--disable-dev-shm-usage')
    
    # Use eager page load strategy
    options.page_load_strategy = 'eager'
    
    # Initialize the driver
    driver = webdriver.Chrome(options=options)
    
    # Set timeout
    driver.set_page_load_timeout(30)
    
    # Set additional wait time for headless mode
    if headless:
        # Add a general wait after page loads in headless mode
        driver.execute_script("return document.readyState") 
    
    return driver

def safe_get_url(driver, url, max_retries=5, retry_delay=30):
    """Safely navigate to URL with retry for network errors."""
    for attempt in range(max_retries):
        try:
            driver.get(url)
            return True
        except Exception as e:
            if "err_internet_disconnected" in str(e).lower():
                logging.warning(f"Internet disconnected. Waiting {retry_delay} seconds before retry... ({attempt+1}/{max_retries})")
                time.sleep(retry_delay)
            else:
                logging.error(f"Error navigating to {url}: {e}")
                time.sleep(retry_delay)
    
    logging.error(f"Failed to navigate to {url} after {max_retries} attempts")
    return False

def wait_for_dropdown_options(driver, dropdown_id, timeout=10):
    """Wait for dropdown to have options beyond the default one."""
    try:
        WebDriverWait(driver, timeout).until(
            lambda d: len(Select(d.find_element(By.ID, dropdown_id)).options) > 1
        )
        return True
    except TimeoutException:
        logging.error(f"Timeout waiting for options in dropdown {dropdown_id}")
        return False

def fill_form(driver, result_data, iteration=1):
    """Fill the form with predefined answers based on the iteration and store the results in result_data."""
    logging.info(f"Starting form filling - Iteration {iteration}")
    
    # Wait for form to be loaded
    WebDriverWait(driver, 10).until(
        EC.presence_of_element_located((By.CLASS_NAME, "form-box"))
    )
    
    # Get screen condition based on iteration
    screen_condition_map = {
        1: {"thai": "หน้าจอไม่มีรอย", "english": "No scratches", "condition": "Flawless"},
        2: {"thai": "หน้าจอมีรอยบางๆ", "english": "There are visible marks but they are not clear.", "condition": "Good"},
        3: {"thai": "หน้าจอมีรอยแตกชำรุด", "english": "There are visible scratches.", "condition": "Damaged"}
    }
    
    screen_condition = screen_condition_map.get(iteration, screen_condition_map[1])
    
    # Store condition in result data
    result_data["Condition"] = screen_condition["condition"]
    
    # Define question-answer pairs based on actual question text (not section numbers)
    # The key is a list of keywords that might appear in the question
    # The value is a list of preferred answers in order of preference
    question_answers = {
        # Model question
        "Model": ["เครื่องไทย TH"],
        
        # Insurance question (ประกัน = insurance)
        "ประกัน": ["หมดประกัน"],  # Out of warranty
        
        # Device condition question (สภาพตัวเครื่อง = device condition)
        "สภาพตัวเครื่อง": ["ไม่มีรอยขีดข่วน"],  # No scratches
        
        # Screen condition question (สภาพหน้าจอ = screen condition)
        "สภาพหน้าจอ": [screen_condition["thai"]],
        
        # Screen display question (การแสดงภาพหน้าจอ = screen display)
        "การแสดงภาพหน้าจอ": ["แสดงภาพหน้าจอปกติ"],  # Normal screen display
        
        # Battery health question (สุขภาพแบตเตอรี่ = battery health)
        "สุขภาพแบตเตอรี่": ["แบตเตอรี่ มากกว่า 80%"],  # Battery more than 80%
        
        # Accessories question (อุปกรณ์เสริม = accessories)
        "อุปกรณ์เสริม": ["ไม่มีกล่อง", "อุปกรณ์ไม่ครบ", "มีกล่อง / อุปกรณ์ครบ", "อุปกรณ์ครบ"],
        
        # Usage problems question (ปัญหาด้านการใช้งาน = usage problems)
        "ปัญหาด้านการใช้งาน": ["ไม่มีปัญหา"]  # No problems
    }
    
    # Helper functions
    def is_section_expanded(section_header):
        # If the collapse class is NOT present or contains "in", section is expanded
        try:
            collapse_div = section_header.find_element(By.XPATH, "following-sibling::div")
            return "in" in collapse_div.get_attribute("class")
        except Exception as e:
            logging.error(f"Error checking if section is expanded: {e}")
            return False
    
    def ensure_section_expanded(section_header):
        if not is_section_expanded(section_header):
            try:
                driver.execute_script("arguments[0].scrollIntoView({block: 'center'});", section_header)
                time.sleep(0.5)
                driver.execute_script("arguments[0].click();", section_header)
                time.sleep(1)
                section_title = section_header.find_element(By.CSS_SELECTOR, ".form-title").text
                logging.info(f"Expanded section {section_title}")
            except Exception as e:
                logging.error(f"Failed to expand section: {e}")
    
    def select_option_in_section(section_header, option_text):
        """Try to select a specific option in a section"""
        ensure_section_expanded(section_header)
        
        # Find the form-checkbox div
        try:
            form_body = section_header.find_element(By.XPATH, "following-sibling::div")
            form_checkbox = form_body.find_element(By.CSS_SELECTOR, ".form-checkbox")
            
            # Find all label elements in this section
            labels = form_checkbox.find_elements(By.CSS_SELECTOR, ".checkbox-label")
            if not labels:
                logging.warning(f"Warning: No options found in section")
                return False
            
            # Log available options for debugging
            label_texts = [label.text.strip() for label in labels]
            logging.info(f"Available options: {label_texts}")
            
            # Try to find and select the exact option
            for label in labels:
                label_text = label.text.strip()
                if label_text == option_text:
                    # Get the associated input
                    input_id = label.get_attribute("for")
                    try:
                        input_element = driver.find_element(By.ID, input_id)
                        
                        # Scroll to element
                        driver.execute_script("arguments[0].scrollIntoView({block: 'center'});", input_element)
                        time.sleep(0.5)
                        
                        # Try clicking the input with JavaScript
                        driver.execute_script("arguments[0].click(); arguments[0].checked = true;", input_element)
                        logging.info(f"Selected: {label_text}")
                        time.sleep(1)
                        
                        # Verify selection
                        try:
                            checked = driver.execute_script("return arguments[0].checked;", input_element)
                            if checked:
                                logging.info(f"Confirmed selection: {label_text} is checked")
                            else:
                                logging.warning(f"Selection verification failed, trying again with direct click")
                                try:
                                    # Try direct click as fallback
                                    input_element.click()
                                    time.sleep(0.5)
                                except Exception as e:
                                    logging.error(f"Direct click failed: {e}")
                        except Exception as e:
                            logging.error(f"Error verifying selection: {e}")
                        
                        return True
                    except Exception as e:
                        logging.error(f"Error selecting {label_text}: {e}")
            
            # If we got here, the exact option wasn't found
            logging.warning(f"Option '{option_text}' not found in available options")
            
            # Fallback: If specified option not found, select first option
            try:
                first_label = labels[0]
                first_label_text = first_label.text.strip()
                input_id = first_label.get_attribute("for")
                input_element = driver.find_element(By.ID, input_id)
                driver.execute_script("arguments[0].scrollIntoView({block: 'center'});", input_element)
                driver.execute_script("arguments[0].click(); arguments[0].checked = true;", input_element)
                logging.info(f"Fallback: Selected first option: {first_label_text}")
                time.sleep(1)
                return True
            except Exception as e:
                logging.error(f"Error selecting fallback option: {e}")
                return False
                
        except Exception as e:
            logging.error(f"Error in select_option_in_section: {e}")
            return False

    def handle_usage_problems_section(section_header):
        """Special handling for Usage Problems section"""
        ensure_section_expanded(section_header)
        
        try:
            # Get all available options
            form_body = section_header.find_element(By.XPATH, "following-sibling::div")
            labels = form_body.find_elements(By.CSS_SELECTOR, ".checkbox-label")
            
            # Log available options
            label_texts = [label.text.strip() for label in labels]
            logging.info(f"Usage problem options: {label_texts}")
            
            # Look for "No problem" option
            no_problem_found = False
            
            # Try multiple approaches
            
            # 1. Try by ID first
            try:
                no_problem = driver.find_element(By.CSS_SELECTOR, "input#optional-no")
                driver.execute_script("arguments[0].scrollIntoView({block: 'center'});", no_problem)
                time.sleep(0.5)
                driver.execute_script("arguments[0].click(); arguments[0].checked = true;", no_problem)
                logging.info("Selected: ไม่มีปัญหา (No problem) by ID")
                time.sleep(1)
                
                # Verify
                checked = driver.execute_script("return arguments[0].checked;", no_problem)
                if checked:
                    logging.info("No problem option successfully checked")
                    return True
                else:
                    logging.warning("No problem option not checked, trying alternative approach")
            except Exception as e:
                logging.info(f"Could not find No problem option by ID: {e}")
            
            # 2. Try by text
            for label in labels:
                if "ไม่มีปัญหา" in label.text:
                    input_id = label.get_attribute("for")
                    input_element = driver.find_element(By.ID, input_id)
                    driver.execute_script("arguments[0].scrollIntoView({block: 'center'});", input_element)
                    time.sleep(0.5)
                    driver.execute_script("arguments[0].click(); arguments[0].checked = true;", input_element)
                    logging.info(f"Selected: {label.text} (No problem)")
                    time.sleep(1)
                    no_problem_found = True
                    return True
            
            # 3. If not found, try the last option (often "No problem")
            if not no_problem_found and labels:
                last_label = labels[-1]
                logging.info(f"Trying last option: {last_label.text}")
                
                input_id = last_label.get_attribute("for")
                input_element = driver.find_element(By.ID, input_id)
                driver.execute_script("arguments[0].scrollIntoView({block: 'center'});", input_element)
                time.sleep(0.5)
                driver.execute_script("arguments[0].click(); arguments[0].checked = true;", input_element)
                logging.info(f"Selected last option (likely No problem): {last_label.text}")
                time.sleep(1)
                return True
            
            return False
        except Exception as e:
            logging.error(f"Error handling usage problems section: {e}")
            return False
    
    # Get all form section headers
    try:
        form_sections = WebDriverWait(driver, 10).until(
            EC.presence_of_all_elements_located((By.CSS_SELECTOR, ".form-box-header"))
        )
        logging.info(f"Found {len(form_sections)} form sections")
    except Exception as e:
        logging.error(f"Error finding form sections: {e}")
        form_sections = []
    
    # Process each form section
    for section_header in form_sections:
        try:
            section_title = section_header.find_element(By.CSS_SELECTOR, ".form-title").text.strip()
            logging.info(f"Processing section: {section_title}")
            
            # Special handling for Usage Problems section
            if "ปัญหาด้านการใช้งาน" in section_title:
                handle_usage_problems_section(section_header)
                continue
            
            # For all other sections, match by keywords in question text
            matched = False
            
            # Try to match the section title with our question keywords
            for question_key, answer_options in question_answers.items():
                if question_key in section_title:
                    logging.info(f"Matched section '{section_title}' with keyword '{question_key}'")
                    
                    # Try each answer option in order until one succeeds
                    for answer_option in answer_options:
                        if select_option_in_section(section_header, answer_option):
                            matched = True
                            break
                    
                    break  # Break out of the question loop if we found a match
            
            # If no match found, use fallback
            if not matched:
                logging.info(f"No match found for section '{section_title}', using fallback")
                ensure_section_expanded(section_header)
                try:
                    form_body = section_header.find_element(By.XPATH, "following-sibling::div")
                    form_checkbox = form_body.find_element(By.CSS_SELECTOR, ".form-checkbox")
                    labels = form_checkbox.find_elements(By.CSS_SELECTOR, ".checkbox-label")
                    if labels:
                        first_label = labels[0]
                        input_id = first_label.get_attribute("for")
                        input_element = driver.find_element(By.ID, input_id)
                        driver.execute_script("arguments[0].scrollIntoView({block: 'center'});", input_element)
                        time.sleep(0.5)
                        driver.execute_script("arguments[0].click(); arguments[0].checked = true;", input_element)
                        logging.info(f"Selected fallback option: {first_label.text}")
                        time.sleep(1)
                except Exception as e:
                    logging.error(f"Error selecting fallback option: {e}")
        except Exception as e:
            logging.error(f"Error processing section: {e}")
    
    # Submit the form
    try:
        # Find the evaluate button (ประเมินราคา)
        evaluate_button = WebDriverWait(driver, 10).until(
            EC.presence_of_element_located((By.CSS_SELECTOR, "a.btn.btn-yello"))
        )
        
        # Scroll to make button visible
        driver.execute_script("arguments[0].scrollIntoView({block: 'center'});", evaluate_button)
        time.sleep(1)
        
        # Remove 'disabled' class if it exists
        driver.execute_script("arguments[0].classList.remove('disabled');", evaluate_button)
        time.sleep(1)
        
        # Click the button
        driver.execute_script("arguments[0].click();", evaluate_button)
        logging.info("Clicked evaluate button")
        time.sleep(5)
        
        # Wait for results page
        try:
            WebDriverWait(driver, 10).until(
                EC.presence_of_element_located((By.CSS_SELECTOR, ".col-xs-12.p-0.bg-white"))
            )
            
            # Extract price
            try:
                price_element = driver.find_element(By.CSS_SELECTOR, "span.red")
                price_text = price_element.text.strip()
                # Clean up the price (remove non-numeric characters except decimal point)
                price_clean = re.sub(r'[^\d.]', '', price_text)
                
                logging.info(f"Extracted price: {price_clean}")
                result_data["Value"] = price_clean
                
                # Also extract device info for better logging
                try:
                    device_info = driver.find_element(By.CSS_SELECTOR, ".product-detail h5").text.strip()
                    logging.info(f"Device: {device_info}, Price: {price_clean}")
                except NoSuchElementException:
                    pass
                
                return True
                
            except NoSuchElementException:
                logging.warning("Could not find price element")
                result_data["Value"] = "NA"
                return False
                
        except TimeoutException:
            logging.error("Timed out waiting for results page")
            result_data["Value"] = "NA"
            return False
            
    except TimeoutException:
        logging.error("Timed out waiting for evaluate button")
        result_data["Value"] = "NA"
        return False
    except Exception as e:
        logging.error(f"Error submitting form: {e}")
        result_data["Value"] = "NA"
        return False

# New function to save results to Excel, matching the format from TH_RV_Source1
def save_to_excel(data, output_file):
    """Save the extracted trade-in data to an Excel file."""
    # Ensure output directory exists
    os.makedirs(os.path.dirname(output_file) if os.path.dirname(output_file) else ".", exist_ok=True)
    
    if os.path.exists(output_file):
        try:
            workbook = openpyxl.load_workbook(output_file)
            sheet = workbook.active
        except:
            workbook = openpyxl.Workbook()
            sheet = workbook.active
            # Use the new column headers
            headers = ["Country", "Device Type", "Brand", "Model", "Capacity", "Color", 
                      "Launch RRP", "Condition", "Value Type", "Currency", "Value", 
                      "Source", "Updated on", "Updated by", "Comments"]
            sheet.append(headers)
    else:
        workbook = openpyxl.Workbook()
        sheet = workbook.active
        # Use the new column headers
        headers = ["Country", "Device Type", "Brand", "Model", "Capacity", "Color", 
                  "Launch RRP", "Condition", "Value Type", "Currency", "Value", 
                  "Source", "Updated on", "Updated by", "Comments"]
        sheet.append(headers)

    # Add new row with all the required columns
    row_data = [
        data["Country"],
        data["Device Type"],
        data["Brand"],
        data["Model"],
        data["Capacity"],
        data["Color"],
        data["Launch RRP"],
        data["Condition"],
        data["Value Type"],
        data["Currency"],
        data["Value"],
        data["Source"],
        data["Updated on"],
        data["Updated by"],
        data["Comments"]
    ]
    sheet.append(row_data)

    try:
        workbook.save(output_file)
        logging.info(f"Saved data to {output_file}")
    except Exception as e:
        logging.error(f"Error saving Excel file: {e}")
        alt_file_name = os.path.join(os.path.dirname(output_file), f"TH_RV_Source2.xlsx")
        try:
            workbook.save(alt_file_name)
            logging.info(f"Saved to {alt_file_name}")
        except:
            logging.error("Could not save Excel file")

# Helper function to determine device type
def detect_device_type(device_name):
    """Determine if the device is a tablet or smartphone based on the name."""
    device_name_lower = device_name.lower()
    
    # Check for tablet keywords
    if any(keyword in device_name_lower for keyword in ["ipad", "tab", "tablet", "galaxy tab"]):
        return "Tablet"
    elif any(keyword in device_name_lower for keyword in ["watch"]):
        return "SmartWatch"
    else:
        # Default to Smartphone for other devices
        return "SmartPhone"

def main_navigation(output_file=None, log_file=None, iterations=3):
    """Main function to navigate through brands, models, and storage options."""
    # Configure logging
    log_file = setup_logging(log_file)
    
    # Use default output path if not specified
    if output_file is None:
        # Check for environment variable first
        output_dir = os.environ.get("OUTPUT_DIR", "output")
        os.makedirs(output_dir, exist_ok=True)
        output_file = os.path.join(output_dir, "TH_RV_Source2.xlsx")
    
    logging.info(f"Will save results to: {output_file}")
    
    # Target brand values to look for with their display names
    target_brand_map = {
        "104": {"name": "Apple", "display": "iPhone"},
        "128": {"name": "Apple", "display": "Apple iPad"},
        "148": {"name": "Samsung", "display": "Samsung Tab"},
        "105": {"name": "Samsung", "display": "Samsung"}
    }
    
    # Setup driver
    driver = setup_driver(headless=False)  # Set to False to see the browser
    
    try:
        # Load the initial page
        if not safe_get_url(driver, "https://www.yellobe.com/", max_retries=10, retry_delay=30):
            logging.error("Failed to load initial page, exiting")
            driver.quit()
            return
        
        time.sleep(4)  # Wait for page to load
        
        # Process each brand
        for brand_value, brand_info in target_brand_map.items():
            brand_display = brand_info["display"]
            brand_name = brand_info["name"]
            
            logging.info(f"\n========== Navigating to {brand_display} (value={brand_value}) ==========\n")
            
            # Load the page fresh for each brand
            if not safe_get_url(driver, "https://www.yellobe.com/", max_retries=10, retry_delay=30):
                logging.error(f"Failed to load page for brand {brand_display}, skipping")
                continue
                
            time.sleep(4)  # Wait for page to load
            
            # Find and select brand
            try:
                brand_dropdown = Select(driver.find_element(By.ID, "brand"))
                logging.info("Found brand dropdown")
                
                # Print brand options for debugging
                logging.info("Brand options:")
                for option in brand_dropdown.options:
                    logging.info(f" - {option.text} (value: {option.get_attribute('value')})")
                
                # Select the brand
                brand_dropdown.select_by_value(brand_value)
                logging.info(f"Selected brand: {brand_display}")
                time.sleep(3)
            except Exception as e:
                if "err_internet_disconnected" in str(e).lower():
                    logging.warning(f"Internet disconnected, waiting 30 seconds before retrying...")
                    time.sleep(30)
                    continue
                else:
                    logging.error(f"Error selecting brand: {e}")
                    continue
            
            # Wait for model dropdown to populate
            if not wait_for_dropdown_options(driver, "seri"):
                logging.error(f"Model dropdown did not populate for brand {brand_display}")
                continue
            
            # Find model dropdown after brand selection
            try:
                model_dropdown = Select(driver.find_element(By.ID, "seri"))
                
                # Store model options for processing
                model_options = []
                for option in model_dropdown.options:
                    if option.get_attribute('value'):  # Skip empty placeholder
                        model_options.append({
                            "value": option.get_attribute('value'),
                            "text": option.text
                        })
                
                logging.info("Model options:")
                for model in model_options:
                    logging.info(f" - {model['text']} (value: {model['value']})")
            except Exception as e:
                if "err_internet_disconnected" in str(e).lower():
                    logging.warning(f"Internet disconnected, waiting 30 seconds before retrying...")
                    time.sleep(30)
                    continue
                else:
                    logging.error(f"Error getting model options: {e}")
                    continue
            
            # Process each model one by one
            for model_data in model_options:
                model_value = model_data["value"]
                model_text = model_data["text"]
                
                logging.info(f"\n=== Navigating to model: {model_text} ===")
                
                # Load fresh page for each model to avoid stale elements
                if not safe_get_url(driver, "https://www.yellobe.com/", max_retries=10, retry_delay=30):
                    logging.error(f"Failed to load page for model {model_text}, skipping")
                    continue
                
                time.sleep(4)
                
                try:
                    # Select brand again
                    brand_dropdown = Select(driver.find_element(By.ID, "brand"))
                    brand_dropdown.select_by_value(brand_value)
                    time.sleep(3)
                    
                    # Wait for model dropdown to populate
                    if not wait_for_dropdown_options(driver, "seri"):
                        logging.error(f"Model dropdown did not populate for brand {brand_display}")
                        continue
                    
                    # Select model
                    model_dropdown = Select(driver.find_element(By.ID, "seri"))
                    model_dropdown.select_by_value(model_value)
                    logging.info(f"Selected model: {model_text}")
                    time.sleep(3)
                except Exception as e:
                    if "err_internet_disconnected" in str(e).lower():
                        logging.warning(f"Internet disconnected, waiting 30 seconds before retrying...")
                        time.sleep(30)
                        continue
                    else:
                        logging.error(f"Error selecting model: {e}")
                        continue
                
                # Determine device type
                device_type = detect_device_type(model_text)
                logging.info(f"Detected device type: {device_type}")
                
                # Wait for storage dropdown to populate
                if not wait_for_dropdown_options(driver, "size_id"):
                    logging.warning(f"Storage dropdown did not populate for model {model_text}, continuing anyway")
                
                # Find storage dropdown
                try:
                    storage_dropdown = Select(driver.find_element(By.ID, "size_id"))
                    
                    # Store storage options for processing
                    storage_options = []
                    for option in storage_dropdown.options:
                        if option.get_attribute('value'):  # Skip empty placeholder
                            storage_options.append({
                                "value": option.get_attribute('value'),
                                "text": option.text
                            })
                    
                    if not storage_options:
                        logging.warning(f"No storage options found for {model_text}")
                        continue
                    
                    logging.info("Storage options:")
                    for storage in storage_options:
                        logging.info(f" - {storage['text']} (value: {storage['value']})")
                    
                except Exception as e:
                    if "err_internet_disconnected" in str(e).lower():
                        logging.warning(f"Internet disconnected, waiting 30 seconds before retrying...")
                        time.sleep(30)
                        continue
                    else:
                        logging.error(f"Error with storage dropdown for {model_text}: {e}")
                        continue
                
                # Process each storage option one by one
                for storage_index, storage_data in enumerate(storage_options):
                    storage_value = storage_data["value"]
                    storage_text = storage_data["text"]
                    
                    logging.info(f"\n--- Selecting capacity: {storage_text} ---")
                    
                    # Process this storage multiple times for different screen condition iterations
                    for iteration in range(1, iterations + 1):
                        logging.info(f"\n--- Iteration {iteration} for capacity: {storage_text} ---")
                        
                        # Load the page fresh
                        if not safe_get_url(driver, "https://www.yellobe.com/", max_retries=10, retry_delay=30):
                            logging.error(f"Failed to load page for storage {storage_text}, iteration {iteration}, skipping")
                            continue
                        
                        time.sleep(4)
                        
                        try:
                            # Re-select brand
                            brand_dropdown = Select(driver.find_element(By.ID, "brand"))
                            brand_dropdown.select_by_value(brand_value)
                            time.sleep(3)
                            
                            # Wait for model dropdown to populate
                            if not wait_for_dropdown_options(driver, "seri"):
                                logging.error(f"Model dropdown did not populate for brand {brand_display}")
                                continue
                            
                            # Re-select model
                            model_dropdown = Select(driver.find_element(By.ID, "seri"))
                            model_dropdown.select_by_value(model_value)
                            time.sleep(3)
                            
                            # Wait for storage dropdown to populate
                            if not wait_for_dropdown_options(driver, "size_id"):
                                logging.warning(f"Storage dropdown did not repopulate, skipping this iteration")
                                continue
                            
                            # Get a fresh handle to the storage dropdown
                            storage_dropdown = Select(driver.find_element(By.ID, "size_id"))
                            
                            # Select storage
                            storage_dropdown.select_by_value(storage_value)
                            logging.info(f"Selected storage: {storage_text}")
                            time.sleep(1)
                            
                            # Find and click submit button
                            submit_button = driver.find_element(By.CSS_SELECTOR, "button.btn-submit")
                            logging.info("Found submit button, clicking...")
                            
                            # Scroll to make button visible
                            driver.execute_script("arguments[0].scrollIntoView({block: 'center'});", submit_button)
                            time.sleep(1)
                            
                            # Click submit
                            submit_button.click()
                            logging.info("Clicked submit button")
                            time.sleep(5)
                            
                            # Create data dictionary for this configuration
                            result_data = {
                                "Country": "Thailand",
                                "Device Type": device_type,
                                "Brand": brand_name,
                                "Model": model_text,
                                "Capacity": storage_text,
                                "Color": "",
                                "Launch RRP": "",
                                "Condition": "",  # Will be set during fill_form
                                "Value Type": "Trade-in",
                                "Currency": "THB",
                                "Value": "",  # Will be set during fill_form
                                "Source": "TH_RV_Source2",
                                "Updated on": datetime.now().strftime("%Y-%m-%d"),
                                "Updated by": "",
                                "Comments": ""
                            }
                            
                            # Fill the form on the next page
                            if fill_form(driver, result_data, iteration):
                                logging.info(f"Successfully extracted trade-in value: {result_data['Value']} THB")
                            else:
                                logging.warning(f"Failed to extract trade-in value for {storage_text}")
                                # Value and Condition should already be set in fill_form
                            
                            # Save the result to Excel regardless of success
                            save_to_excel(result_data, output_file)
                            
                        except Exception as e:
                            if "err_internet_disconnected" in str(e).lower():
                                logging.warning(f"Internet disconnected during processing of {storage_text}. Waiting 30 seconds before retrying...")
                                time.sleep(30)
                                # Continue with the next iteration
                                continue
                            else:
                                logging.error(f"Error processing storage {storage_text}: {e}")
    
    except Exception as e:
        if "err_internet_disconnected" in str(e).lower():
            logging.warning(f"Internet disconnected. Waiting 30 seconds before retrying the entire process...")
            time.sleep(30)
            # Recursive call to retry the entire process
            main_navigation(output_file, log_file, iterations)
        else:
            logging.error(f"Error in main navigation: {e}")
            import traceback
            logging.error(traceback.format_exc())
    
    finally:
        driver.quit()
        logging.info("Browser closed. Navigation complete.")

if __name__ == "__main__":
    parser = argparse.ArgumentParser(description='Navigate through dropdowns and fill forms on Thailand Yellobe website')
    parser.add_argument('-o', '--output', type=str, help='Output Excel file path', default=None)
    parser.add_argument('-l', '--log', type=str, help='Log file path', default=None)
    parser.add_argument('-i', '--iterations', type=int, help='Number of iterations with different screen conditions', default=3)
    args = parser.parse_args()
    
    main_navigation(args.output, args.log, args.iterations)