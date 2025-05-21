from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import Select, WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.chrome.service import Service
from selenium.common.exceptions import TimeoutException, StaleElementReferenceException, ElementNotInteractableException
from webdriver_manager.chrome import ChromeDriverManager
import time
import openpyxl
from datetime import datetime
import os
import re

# Setup Chrome driver with appropriate options
def setup_driver(headless=True):
    """Set up the Chrome WebDriver with appropriate options."""
    options = webdriver.ChromeOptions()
    
    # Only enable headless mode if specified
    if headless:
        options.add_argument('--headless')
    
    # Core options for stability
    options.add_argument('--no-sandbox')
    options.add_argument('--disable-dev-shm-usage')
    options.add_argument('--disable-gpu')
    options.add_argument('--start-maximized')
    options.add_argument('--window-size=1920,1080')
    
    # Performance options
    options.add_argument('--disable-extensions')
    options.add_argument('--disable-infobars')
    options.add_argument('--disable-logging')
    options.add_argument('--disable-notifications')
    
    # User agent
    options.add_argument('--user-agent=Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/91.0.4472.124 Safari/537.36')
    
    return webdriver.Chrome(options=options)

# Function to select dropdowns
def select_dropdowns(driver, brand, device_type):
    print(f"Selecting dropdowns for {brand} {device_type}...")
    
    # Wait for page to fully load
    WebDriverWait(driver, 20).until(
        EC.presence_of_element_located((By.ID, "select-brand"))
    )
    time.sleep(2)  # Additional wait to ensure JavaScript is fully loaded
    
    try:
        # Select brand
        brand_dropdown = WebDriverWait(driver, 10).until(
            EC.element_to_be_clickable((By.ID, "select-brand"))
        )
        Select(brand_dropdown).select_by_value(brand)
        print(f"Selected brand: {brand}")
        time.sleep(2)  # Wait for the dropdown to update

        # Select type
        type_dropdown = WebDriverWait(driver, 10).until(
            EC.element_to_be_clickable((By.ID, "select-type"))
        )
        Select(type_dropdown).select_by_value(device_type)
        print(f"Selected type: {device_type}")
        time.sleep(2)  # Wait for model options to load
        
        # Check for model dropdown presence
        WebDriverWait(driver, 15).until(
            EC.presence_of_element_located((By.CSS_SELECTOR, "#select-model option:not([value=''])"))
        )
        
        # Get all model options
        model_dropdown = WebDriverWait(driver, 10).until(
            EC.element_to_be_clickable((By.ID, "select-model"))
        )
        model_select = Select(model_dropdown)
        models = []
        
        # Skip the first option which is typically empty or a placeholder
        for index, option in enumerate(model_select.options):
            if index > 0 and option.get_attribute("value"):
                models.append({
                    "index": index,
                    "value": option.get_attribute("value"),
                    "text": option.text
                })
        
        return models
                
    except Exception as e:
        print(f"Error in select_dropdowns: {e}")
        return []

# Function to process a single model
def process_model(driver, brand, device_type, model, country="TH", output_file=None):
    print(f"\nProcessing model: {model['text']}")
    
    try:
        # Select model from dropdown
        model_dropdown = WebDriverWait(driver, 10).until(
            EC.element_to_be_clickable((By.ID, "select-model"))
        )
        model_select = Select(model_dropdown)
        model_select.select_by_index(model['index'])
        print(f"Selected model: {model['text']}")
        time.sleep(2)  # Wait for storage options to load
        
        # Get all storage options
        WebDriverWait(driver, 15).until(
            EC.presence_of_element_located((By.CSS_SELECTOR, "#select-storage option:not([value=''])"))
        )
        
        storage_dropdown = WebDriverWait(driver, 10).until(
            EC.element_to_be_clickable((By.ID, "select-storage"))
        )
        storage_select = Select(storage_dropdown)
        
        # Skip the first option which is typically empty or a placeholder
        storages = []
        for index, option in enumerate(storage_select.options):
            if index > 0 and option.get_attribute("value"):
                storages.append({
                    "index": index,
                    "value": option.get_attribute("value"),
                    "text": option.text
                })
        
        all_storage_results = []
        
        # Process each storage option
        for i, storage in enumerate(storages):
            print(f"Processing storage: {storage['text']} ({i+1}/{len(storages)})")
            
            # For each storage, we start fresh with a new page load
            # This ensures dropdown states are consistent
            driver.get("https://www.kaitorasap.co.th/sale-phone")
            time.sleep(2)
            
            # Wait for brand dropdown
            WebDriverWait(driver, 20).until(
                EC.presence_of_element_located((By.ID, "select-brand"))
            )
            
            # Reselect brand
            brand_dropdown = WebDriverWait(driver, 10).until(
                EC.element_to_be_clickable((By.ID, "select-brand"))
            )
            Select(brand_dropdown).select_by_value(brand)
            time.sleep(1)
            
            # Reselect type
            type_dropdown = WebDriverWait(driver, 10).until(
                EC.element_to_be_clickable((By.ID, "select-type"))
            )
            Select(type_dropdown).select_by_value(device_type)
            time.sleep(1)
            
            # Reselect model
            model_dropdown = WebDriverWait(driver, 10).until(
                EC.element_to_be_clickable((By.ID, "select-model"))
            )
            model_select = Select(model_dropdown)
            model_select.select_by_index(model['index'])
            time.sleep(1)
            
            # Select the storage
            storage_dropdown = WebDriverWait(driver, 10).until(
                EC.element_to_be_clickable((By.ID, "select-storage"))
            )
            storage_select = Select(storage_dropdown)
            storage_select.select_by_index(storage['index'])
            time.sleep(1)
            
            # For Apple devices, select country
            if brand == "APPLE":
                try:
                    country_dropdown = WebDriverWait(driver, 10).until(
                        EC.element_to_be_clickable((By.ID, "select-model-country"))
                    )
                    Select(country_dropdown).select_by_value(country)
                    print(f"Selected country: {country}")
                    time.sleep(1)
                except Exception as e:
                    print(f"Warning: Could not select country. Error: {e}")
            
            # Click Confirm button
            try:
                confirm_button = WebDriverWait(driver, 15).until(
                    EC.element_to_be_clickable((By.CLASS_NAME, "bt-conf-select-model"))
                )
                try:
                    confirm_button.click()
                except:
                    driver.execute_script("arguments[0].click();", confirm_button)
            except Exception as e:
                print(f"Error clicking confirm button: {e}")
                continue  # Skip this storage and try the next one
            
            # Wait for condition section to load
            try:
                WebDriverWait(driver, 20).until(
                    EC.visibility_of_element_located((By.ID, "bg-select-detail-model"))
                )
                time.sleep(2)  # Allow page to stabilize
                
                # Process each condition
                conditions = [
                    {"name": "Flawless", "no_problem": True, "scratch_id": "scratch_a", "damaged_id": None},
                    {"name": "Good", "no_problem": True, "scratch_id": "scratch_b", "damaged_id": None},
                    {"name": "Damaged", "no_problem": False, "scratch_id": "scratch_a", "damaged_id": "over2"}
                ]
                
                storage_results = []
                
                for condition in conditions:
                    try:
                        price = get_price_for_condition(
                            driver, 
                            condition["no_problem"], 
                            condition["scratch_id"], 
                            condition["damaged_id"]
                        )
                        
                        # Create a result entry
                        result = {
                            "Model": model['text'],
                            "Storage": storage['text'],
                            "Condition": condition["name"],
                            "Price": price if price else "NA",
                            "Brand": brand,
                            "Device_Type": device_type
                        }
                        
                        storage_results.append(result)
                        all_storage_results.append(result)
                        print(f"{condition['name']}: {price if price else 'NA'}")
                        
                        # Save after each condition is processed
                        if output_file and len(storage_results) > 0:
                            save_to_excel([result], output_file)
                        
                    except Exception as e:
                        print(f"Error processing condition {condition['name']}: {e}")
                        # Create an entry with NA for price
                        result = {
                            "Model": model['text'],
                            "Storage": storage['text'],
                            "Condition": condition["name"],
                            "Price": "NA",
                            "Brand": brand,
                            "Device_Type": device_type
                        }
                        storage_results.append(result)
                        all_storage_results.append(result)
                        
                        # Save after each condition is processed, even with NA
                        if output_file:
                            save_to_excel([result], output_file)
                
            except Exception as e:
                print(f"Error in condition section: {e}")
                # Create an entry with NA for all conditions
                for condition_name in ["Flawless", "Good", "Damaged"]:
                    result = {
                        "Model": model['text'],
                        "Storage": storage['text'],
                        "Condition": condition_name,
                        "Price": "NA",
                        "Brand": brand,
                        "Device_Type": device_type
                    }
                    all_storage_results.append(result)
                    
                    # Save each NA result
                    if output_file:
                        save_to_excel([result], output_file)
        
        return all_storage_results
        
    except Exception as e:
        print(f"Error processing model {model['text']}: {e}")
        return []

# Function to select conditions and get price
def get_price_for_condition(driver, no_problem=True, scratch_id="scratch_a", damaged_id=None):
    try:
        # Wait for condition section to be visible
        condition_section = WebDriverWait(driver, 15).until(
            EC.visibility_of_element_located((By.ID, "bg-select-detail-model"))
        )
        
        # Select condition checkboxes
        if no_problem:
            no_problem_checkbox = WebDriverWait(driver, 15).until(
                EC.presence_of_element_located((By.ID, "no_problem"))
            )
            
            # Force visibility and click
            driver.execute_script("arguments[0].style.display = 'block'; arguments[0].style.visibility = 'visible';", no_problem_checkbox)
            time.sleep(1)
            
            try:
                no_problem_checkbox.click()
            except:
                driver.execute_script("arguments[0].click();", no_problem_checkbox)
                
        elif damaged_id:
            damaged_checkbox = WebDriverWait(driver, 15).until(
                EC.presence_of_element_located((By.ID, damaged_id))
            )
            
            # Force visibility and click
            driver.execute_script("arguments[0].style.display = 'block'; arguments[0].style.visibility = 'visible';", damaged_checkbox)
            time.sleep(1)
            
            try:
                damaged_checkbox.click()
            except:
                driver.execute_script("arguments[0].click();", damaged_checkbox)
        
        time.sleep(2)  # Wait for scratch section to appear

        # Select scratch radio button via label
        WebDriverWait(driver, 15).until(
            EC.visibility_of_element_located((By.ID, "select-scratch-all"))
        )
        
        scratch_label = WebDriverWait(driver, 15).until(
            EC.element_to_be_clickable((By.CSS_SELECTOR, f"label[for='{scratch_id}']"))
        )
        
        try:
            scratch_label.click()
        except:
            driver.execute_script("arguments[0].click();", scratch_label)
        
        time.sleep(2)  # Wait for accessories section to update

        # Select "No accessories included"
        acc_label = WebDriverWait(driver, 15).until(
            EC.element_to_be_clickable((By.CSS_SELECTOR, "label[for='acc-n']"))
        )
        
        try:
            acc_label.click()
        except:
            driver.execute_script("arguments[0].click();", acc_label)
        
        time.sleep(2)  # Wait for price to update

        # Extract price
        price_element = WebDriverWait(driver, 15).until(
            EC.visibility_of_element_located((By.ID, "price-re"))
        )
        price_text = price_element.text.strip()
        
        # Clean up price (remove commas, currency symbols)
        price_clean = re.sub(r'[^\d.]', '', price_text)
        return price_clean
        
    except Exception as e:
        print(f"Error in get_price_for_condition: {e}")
        return None

# Function to save results to Excel
def save_to_excel(results, output_file):
    """Save the extracted trade-in data to an Excel file."""
    # Ensure output directory exists
    os.makedirs(os.path.dirname(output_file) if os.path.dirname(output_file) else ".", exist_ok=True)
    
    # Check if file exists and load it or create new
    if os.path.exists(output_file):
        try:
            workbook = openpyxl.load_workbook(output_file)
            sheet = workbook.active
        except Exception as e:
            print(f"Error opening existing file: {e}. Creating new file.")
            workbook = openpyxl.Workbook()
            sheet = workbook.active
            # Use the column headers from TH_RV_Source1.py
            headers = ["Country", "Device Type", "Brand", "Model", "Capacity", "Color", 
                      "Launch RRP", "Condition", "Value Type", "Currency", "Value", 
                      "Source", "Updated on", "Updated by", "Comments"]
            sheet.append(headers)
    else:
        workbook = openpyxl.Workbook()
        sheet = workbook.active
        # Use the column headers from TH_RV_Source1.py
        headers = ["Country", "Device Type", "Brand", "Model", "Capacity", "Color", 
                  "Launch RRP", "Condition", "Value Type", "Currency", "Value", 
                  "Source", "Updated on", "Updated by", "Comments"]
        sheet.append(headers)

    # Map device types
    device_type_map = {
        "iPhone": "SmartPhone",
        "iPad": "Tablet",
        "Mobile": "SmartPhone",
        "Tablet": "Tablet"
    }

    # Get the current date
    current_date = datetime.now().strftime("%Y-%m-%d")

    # Add new rows with all the required columns
    for result in results:
        # Map device type
        mapped_device_type = device_type_map.get(result["Device_Type"], "SmartPhone")
        
        row_data = [
            "Thailand",  # Country
            mapped_device_type,  # Device Type
            result["Brand"],  # Brand
            result["Model"],  # Model
            result["Storage"],  # Capacity
            "",  # Color
            "",  # Launch RRP
            result["Condition"],  # Condition
            "Trade-in",  # Value Type
            "THB",  # Currency
            result["Price"],  # Value
            "TH_RV_Source3",  # Source
            current_date,  # Updated on
            "",  # Updated by
            ""  # Comments
        ]
        sheet.append(row_data)

    try:
        workbook.save(output_file)
        print(f"Data saved to {output_file}")
    except Exception as e:
        print(f"Error saving Excel file: {e}")
        # Try to save with a different filename as fallback
        try:
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            alt_file_name = os.path.join(os.path.dirname(output_file), f"TH_RV_Source3_{timestamp}.xlsx")
            workbook.save(alt_file_name)
            print(f"Saved to alternative file: {alt_file_name}")
        except Exception as e2:
            print(f"Could not save to alternative file either: {e2}")
            return False
    
    return True
def main():
    # Define brands and device types to scrape
    brand_device_types = [
        {"brand": "APPLE", "device_type": "iPhone"},
        {"brand": "APPLE", "device_type": "iPad"},
        {"brand": "SAMSUNG", "device_type": "Mobile"},
        {"brand": "SAMSUNG", "device_type": "Tablet"}
    ]
    
    # Output file
    output_dir = os.environ.get("OUTPUT_DIR", "output")
    os.makedirs(output_dir, exist_ok=True)
    output_file = os.path.join(output_dir, "TH_RV_Source3.xlsx")
    
    # Initialize driver
    driver = setup_driver(headless=True)
    
    # Set window size to ensure elements are visible
    driver.set_window_size(1366, 1000)
    
    try:
        for item in brand_device_types:
            brand = item["brand"]
            device_type = item["device_type"]
            
            print(f"\n=== Processing {brand} {device_type} ===\n")
            
            # Load the website
            driver.get("https://www.kaitorasap.co.th/sale-phone")
            time.sleep(3)  # Wait for initial page load
            
            # Select brand and device type, get models
            models = select_dropdowns(driver, brand, device_type)
            print(f"Found {len(models)} models for {brand} {device_type}")
            
            # Process each model
            for model in models:
                try:
                    # Process current model and get results
                    process_model(driver, brand, device_type, model, output_file=output_file)
                    # Note: We don't collect results here anymore since they're saved after each price
                except Exception as e:
                    print(f"Error processing {model['text']}: {e}")
                    continue
    
    except Exception as e:
        print(f"An error occurred in main process: {e}")
    
    finally:
        # Close the browser
        print("Closing browser...")
        driver.quit()
        print("Process completed")

if __name__ == "__main__":
    main()