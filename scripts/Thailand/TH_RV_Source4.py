from selenium import webdriver
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.common.exceptions import TimeoutException, NoSuchElementException, StaleElementReferenceException
from webdriver_manager.chrome import ChromeDriverManager
import time
import openpyxl
from datetime import datetime
import os
import re
import argparse

def setup_driver(headless=True):
    """Set up the Chrome WebDriver with appropriate options."""
    options = webdriver.ChromeOptions()
    
    if headless:
        options.add_argument('--headless')
    
    options.add_argument('--no-sandbox')
    options.add_argument('--disable-dev-shm-usage')
    options.add_argument('--disable-gpu')
    options.add_argument('--start-maximized')
    options.add_argument('--window-size=1920,1080')
    
    options.add_argument('--disable-extensions')
    options.add_argument('--disable-infobars')
    options.add_argument('--disable-logging')
    options.add_argument('--disable-notifications')
    
    options.add_argument('--user-agent=Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/91.0.4472.124 Safari/537.36')
    
    driver = webdriver.Chrome(options=options)
    driver.set_page_load_timeout(60)
    
    return driver

def detect_device_type(device_name):
    """Determine if the device is a tablet or smartphone based on the name."""
    device_name_lower = device_name.lower()
    
    if any(keyword in device_name_lower for keyword in ["ipad", "i Pad", "tab", "tablet", "galaxy tab"]):
        return "Tablet"
    elif any(keyword  in device_name_lower for keyword in ["watch"]):
        return "SmartWatch"
    else:
        return "SmartPhone"

def go_to_next_page(driver, wait):
    """Attempt to navigate to the next page of results."""
    try:
        # Target the "Next" button as the last <li> element's <button> in the pagination <ul>
        next_button = wait.until(
            EC.element_to_be_clickable((By.XPATH, "//ul[contains(@class, 'flex justify-between items-center')]//li[last()]//button"))
        )
        
        # Check if the button is disabled
        if next_button.get_attribute("disabled"):
            print("Next page button is disabled, reached last page")
            return False
        
        # Scroll to the button and click
        driver.execute_script("arguments[0].scrollIntoView({block: 'center'});", next_button)
        time.sleep(0.5)
        next_button.click()
        
        # Wait for the page to load and ensure device cards are present
        wait.until(EC.presence_of_element_located((By.CSS_SELECTOR, "div[data-v-b4522aa0][data-v-545eeb04]")))
        time.sleep(1)  # Additional delay for page stability
        
        print("Successfully navigated to next page")
        return True
        
    except NoSuchElementException:
        print("Next page button not found")
        return False
    except TimeoutException:
        print("Timeout waiting for next page button or page load")
        return False
    except Exception as e:
        print(f"Error navigating to next page: {e}")
        return False

def save_to_excel(data, output_file):
    """Save the extracted trade-in data to an Excel file."""
    os.makedirs(os.path.dirname(output_file) if os.path.dirname(output_file) else ".", exist_ok=True)
    
    if os.path.exists(output_file):
        try:
            workbook = openpyxl.load_workbook(output_file)
            sheet = workbook.active
        except:
            workbook = openpyxl.Workbook()
            sheet = workbook.active
            headers = ["Country", "Device Type", "Brand", "Model", "Capacity", "Color", 
                      "Launch RRP", "Condition", "Value Type", "Currency", "Value", 
                      "Source", "Updated on", "Updated by", "Comments"]
            sheet.append(headers)
    else:
        workbook = openpyxl.Workbook()
        sheet = workbook.active
        headers = ["Country", "Device Type", "Brand", "Model", "Capacity", "Color", 
                  "Launch RRP", "Condition", "Value Type", "Currency", "Value", 
                  "Source", "Updated on", "Updated by", "Comments"]
        sheet.append(headers)

    row_data = [
        data["Country"],
        data["Device Type"],
        data["Brand"],
        data["Model"],
        data["Capacity"],
        data["Color"],
        data["Launch RRP"],
        data["Condition"],
        data["Value Type"],
        data["Currency"],
        data["Value"],
        data["Source"],
        data["Updated on"],
        data["Updated by"],
        data["Comments"]
    ]
    sheet.append(row_data)

    try:
        workbook.save(output_file)
        print(f"Saved data to {output_file}")
    except Exception as e:
        print(f"Error saving Excel file: {e}")
        alt_file_name = os.path.join(os.path.dirname(output_file), f"TH_RV_Source4.xlsx")
        try:
            workbook.save(alt_file_name)
            print(f"Saved to {alt_file_name}")
        except:
            print("Could not save Excel file")

def get_question_section(driver, question_keywords):
    """Find a form section by keywords in multiple languages.
    
    Args:
        driver: The WebDriver instance
        question_keywords: List of possible keywords for the question in different languages
    
    Returns:
        The section element if found, None otherwise
    """
    try:
        # First, get all section headers to check what's actually on the page
        all_headers = driver.find_elements(By.XPATH, "//div[contains(@class, 'box-shadow rounded mb-5')]//div[contains(@class, 'bg-[#EBF3EE]')]")
        header_texts = [header.text.strip() for header in all_headers]
        print(f"Found form headers: {header_texts}")
        
        # Try to find section by partial text match for each keyword
        for keyword in question_keywords:
            # Try partial match to be more tolerant of small text differences
            sections = []
            for header_text in header_texts:
                # Check if keyword is a substantial part of the header text
                if any(word.lower() in header_text.lower() for word in keyword.split()):
                    # Find the corresponding section
                    xpath = f"//div[contains(@class, 'box-shadow rounded mb-5')][.//div[contains(@class, 'bg-[#EBF3EE]') and contains(text(), '{header_text}')]]"
                    matching_sections = driver.find_elements(By.XPATH, xpath)
                    if matching_sections:
                        sections.append(matching_sections[0])
                        print(f"Found question section with keyword '{keyword}' matching header '{header_text}'")
            
            if sections:
                return sections[0]
                
            # Fallback to direct contains if partial match didn't work
            xpath = f"//div[contains(@class, 'box-shadow rounded mb-5')][.//div[contains(@class, 'bg-[#EBF3EE]') and contains(., '{keyword}')]]"
            sections = driver.find_elements(By.XPATH, xpath)
            
            if sections:
                print(f"Found question section with keyword: '{keyword}'")
                return sections[0]
        
        # If we get here, none of the keywords matched
        print(f"Question section with keywords {question_keywords} not found")
        return None
    except Exception as e:
        print(f"Error finding question section with keywords {question_keywords}: {e}")
        return None

def select_option_by_text(driver, section, option_keywords):
    """Select an option by keywords in multiple languages.
    
    Args:
        driver: The WebDriver instance
        section: The section element
        option_keywords: List of possible option texts in different languages
    
    Returns:
        True if option was selected, False otherwise
    """
    if section:
        try:
            # Find buttons in this section
            buttons = section.find_elements(By.TAG_NAME, "button")
            
            # Try each keyword until we find a match
            for keyword in option_keywords:
                for button in buttons:
                    if keyword.lower() in button.text.strip().lower():
                        driver.execute_script("arguments[0].scrollIntoView({block: 'center'});", button)
                        time.sleep(0.1)
                        button.click()
                        print(f"Selected option with keyword '{keyword}'")
                        return True
            
            print(f"No options matching keywords {option_keywords} found in section")
            return False
        except Exception as e:
            print(f"Error selecting option with keywords {option_keywords}: {e}")
            return False
    return False

def select_option_by_index(driver, section, option_index):
    """Select an option by its index (1-based) within a section."""
    if section:
        try:
            # Find buttons in this section
            buttons = section.find_elements(By.TAG_NAME, "button")
            
            if 0 <= option_index - 1 < len(buttons):
                button = buttons[option_index - 1]
                
                # Check if already selected
                button_style = button.get_attribute("style")
                if "bg-primary" not in button_style:
                    driver.execute_script("arguments[0].scrollIntoView({block: 'center'});", button)
                    time.sleep(0.1)
                    button.click()
                    print(f"Selected option {option_index} - '{button.text.strip()}'")
                    return True
                else:
                    print(f"Option {option_index} already selected")
                    return True
            else:
                print(f"Option index {option_index} out of range (only {len(buttons)} options available)")
                return False
        except Exception as e:
            print(f"Error selecting option by index {option_index}: {e}")
            return False
    return False

def fill_form_and_get_price(driver, wait, condition_option):
    """Fill the form with the specified options and get the price."""
    try:
        print(f"Filling form for condition: {condition_option}")
        
        # Define question keywords and corresponding option selections in both English and Thai
        question_options = {
            "capacity": {
                "keywords": ["Phone capacity", "ความจุของโทรศัพท์", "ความจุ"],
                # Capacity is handled separately - this is just a placeholder
                "options": {"option_keywords": ["128 GB"], "option_index": 1}
            },
            "burn_in": {
                "keywords": ["Burn-in problem", "ปัญหาจอเบิร์น", "ปัญหาจอเบริน์", "เบิร์น", "เบริน์"],
                "options": {"option_keywords": ["do not have", "ไม่มี"], "option_index": 3}
            },
            "foldable": {
                "keywords": ["Foldable screen", "มือถือจอพับได้", "จอพับ"],
                "options": {"option_keywords": ["Normal screen", "หน้าจอปกติ"], "option_index": 3}
            },
            "accessories": {
                "keywords": ["Available accessories", "อุปกรณ์เสริมที่มี", "อุปกรณ์เสริม"],
                "options": {"option_keywords": ["No box", "ไม่มีกล่อง"], "option_index": 3}
            },
            "problems": {
                "keywords": ["Phone problems", "ปัญหาของโทรศัพท์", "ปัญหาโทรศัพท์"],
                "options": {
                    "Flawless": {"option_keywords": ["no problem", "ไม่มีปัญหา"], "option_index": 13},
                    "Good": {"option_keywords": ["obvious scratches", "หน้าจอมีรอยชัดเจน", "รอยชัดเจน"], "option_index": 8},
                    "Damaged": {"option_keywords": ["Broken screen", "หน้าจอแตก"], "option_index": 7}
                }
            },
            "service_center": {
                "keywords": ["Service center", "ศูนย์บริการที่ซื้อ", "ศูนย์บริการ"],
                "options": {"option_keywords": ["Thai", "ไทย"], "option_index": 1}
            },
            "condition": {
                "keywords": ["Condition of the phone", "สภาพของโทรศัพท์", "สภาพโทรศัพท์"],
                "options": {"option_keywords": ["No trace", "ไม่มีรอย"], "option_index": 4}
            },
            "icloud": {
                "keywords": ["log out of I cloud", "ออกจากระบบ I cloud", "I cloud", "icloud"],
                "options": {"option_keywords": ["can", "ได้"], "option_index": 1}
            },
            "power_on": {
                "keywords": ["turned on", "สามารถเปิดติด", "เปิดติด"],
                "options": {"option_keywords": ["stuck", "ติด"], "option_index": 1}
            },
            "battery": {
                "keywords": ["Battery health", "สุขภาพแบต", "แบต"],
                "options": {"option_keywords": ["yes", "ใช่"], "option_index": 1}
            },
            "warranty": {
                "keywords": ["warranty for 4 months", "มีประกัน 4 เดือน", "ประกัน"],
                "options": {"option_keywords": ["do not have", "ไม่มี"], "option_index": 2}
            }
        }
        
        # Process all form sections by index instead of assuming a specific order
        all_sections = driver.find_elements(By.XPATH, "//div[contains(@class, 'box-shadow rounded mb-5')]")
        
        print(f"Found {len(all_sections)} question sections on the page")
        
        # Skip capacity as it's handled separately
        for question_key, question_data in question_options.items():
            if question_key == "capacity":
                continue
                
            # Phone problems has different options based on condition
            if question_key == "problems":
                condition_specific_options = question_data["options"][condition_option]
                option_keywords = condition_specific_options["option_keywords"]
                option_index = condition_specific_options["option_index"]
            else:
                option_keywords = question_data["options"]["option_keywords"]
                option_index = question_data["options"]["option_index"]
            
            # Find the section for this question
            section = get_question_section(driver, question_data["keywords"])
            
            if section:
                # Try to select by text first, fall back to index
                if not select_option_by_text(driver, section, option_keywords):
                    select_option_by_index(driver, section, option_index)
            else:
                print(f"Skipping question '{question_key}' as it wasn't found")
        
        # Click "View price offers" button - try both English and Thai
        try:
            # Try first with various text content in both languages
            view_price_buttons = driver.find_elements(By.XPATH, 
                "//button[contains(., 'View price offers') or contains(., 'price offers') or " +
                "contains(., 'view price') or contains(., 'ดูข้อเสนอราคา')]")
            
            if not view_price_buttons:
                # Try with the last button in the form
                view_price_buttons = driver.find_elements(By.XPATH, "//div[contains(@class, 'grid')]/button[last()]")
            
            if view_price_buttons:
                view_price_button = view_price_buttons[-1]  # Take the last matching button
                driver.execute_script("arguments[0].scrollIntoView({block: 'center'});", view_price_button)
                time.sleep(1)
                print(f"Clicking 'View price offers' button: {view_price_button.text.strip()}")
                view_price_button.click()
                time.sleep(2)
            else:
                print("⚠️ Could not find 'View price offers' button")
                return "NA"
        except Exception as e:
            print(f"Error clicking 'View price offers' button: {e}")
            return "NA"
        
        # Wait for price to appear and extract it
        try:
            wait.until(EC.presence_of_element_located((By.XPATH, "//input[@placeholder and @disabled]")))
            time.sleep(1)
            
            price_input = driver.find_element(By.XPATH, "//input[@placeholder and @disabled]")
            price_value = price_input.get_attribute("placeholder").replace(",", "")
            print(f"✅ Extracted price for {condition_option} condition: {price_value} THB")
            
            # Click "Retrospective" button to go back - try both English and Thai
            back_buttons = driver.find_elements(By.XPATH, 
                "//button[contains(., 'retrospective') or contains(., 'back') or " +
                "contains(., 'previous') or contains(., 'ย้อนกลับ')]")
            
            if not back_buttons:
                # Try with the first button in the form
                back_buttons = driver.find_elements(By.XPATH, "//div[contains(@class, 'grid')]/button[1]")
            
            if back_buttons:
                back_button = back_buttons[0]  # Take the first matching button
                driver.execute_script("arguments[0].scrollIntoView({block: 'center'});", back_button)
                time.sleep(1)
                print(f"Clicking 'Back' button: {back_button.text.strip()}")
                back_button.click()
                time.sleep(2)
            else:
                print("⚠️ Could not find 'Back' button, navigating back")
                driver.back()
                time.sleep(2)
            
            # Wait for form to reload
            wait.until(EC.presence_of_element_located((By.XPATH, "//div[contains(@class, 'box-shadow rounded mb-5')]")))
            time.sleep(1)
            
            return price_value
            
        except Exception as e:
            print(f"Error extracting price: {e}")
            # Try to navigate back if we reached this point
            driver.back()
            time.sleep(2)
            return "NA"
        
    except Exception as e:
        print(f"Error in filling form or getting price for {condition_option} condition: {e}")
        import traceback
        traceback.print_exc()
        # Try to navigate back if there was an error
        try:
            driver.back()
            time.sleep(2)
        except:
            pass
        return "NA"

def get_capacity_from_button(button):
    """Extract capacity from button text."""
    capacity_text = button.text.strip()
    # Extract numbers and GB from text like "128 GB"
    match = re.search(r'(\d+)\s*GB', capacity_text, re.IGNORECASE)
    if match:
        return match.group(1) + " GB"
    return capacity_text

def process_device_with_prices(driver, wait, brand_name, device_title, output_file):
    """Process a device and get prices for all conditions and capacities."""
    device_type = detect_device_type(device_title)
    print(f"\n=== Processing device: {device_title} ===")
    
    try:
        # Wait for form sections to load
        wait.until(EC.presence_of_element_located((By.XPATH, "//div[contains(@class, 'box-shadow rounded mb-5')]")))
        time.sleep(2)  # Give page a moment to fully render
        
        # Get capacity section by looking for keywords in multiple languages
        capacity_section = get_question_section(driver, ["Phone capacity", "ความจุของโทรศัพท์"])
        
        if not capacity_section:
            print("⚠️ Could not find capacity section")
            return False
        
        capacity_buttons = capacity_section.find_elements(By.TAG_NAME, "button")
        
        # Store capacity texts to avoid stale element issues
        capacities = []
        for button in capacity_buttons:
            try:
                capacity_text = button.text.strip()
                match = re.search(r'(\d+)\s*GB', capacity_text, re.IGNORECASE)
                if match:
                    capacities.append(match.group(1) + " GB")
                else:
                    capacities.append(capacity_text)
            except Exception as e:
                print(f"Error getting capacity text: {e}")
                capacities.append("Unknown")
        
        print(f"Found {len(capacities)} capacity options: {capacities}")
        
        for capacity_index, capacity in enumerate(capacities):
            print(f"\nProcessing capacity: {capacity}")
            
            # Re-fetch the capacity section and buttons to avoid stale references
            capacity_section = get_question_section(driver, ["Phone capacity", "ความจุของโทรศัพท์"])
            if not capacity_section:
                print("⚠️ Could not find capacity section")
                continue
                
            fresh_capacity_buttons = capacity_section.find_elements(By.TAG_NAME, "button")
            
            if capacity_index < len(fresh_capacity_buttons):
                # Click on this capacity option
                capacity_button = fresh_capacity_buttons[capacity_index]
                driver.execute_script("arguments[0].scrollIntoView({block: 'center'});", capacity_button)
                time.sleep(0.5)
                print(f"Clicking capacity button: {capacity}")
                capacity_button.click()
                time.sleep(1)
            else:
                print(f"⚠️ Capacity button at index {capacity_index} not found")
                continue
            
            # Get prices for each condition
            for condition in ["Flawless", "Good", "Damaged"]:
                price = fill_form_and_get_price(driver, wait, condition)
                
                # Save to Excel
                trade_in_data = {
                    "Country": "Thailand",
                    "Device Type": device_type,
                    "Brand": brand_name,
                    "Model": device_title,
                    "Capacity": capacity,
                    "Color": "",
                    "Launch RRP": "",
                    "Condition": condition,
                    "Value Type": "Trade-in",
                    "Currency": "THB",
                    "Value": price if price else "NA",
                    "Source": "TH_RV_Source4",
                    "Updated on": datetime.now().strftime("%Y-%m-%d"),
                    "Updated by": "",
                    "Comments": ""
                }
                
                save_to_excel(trade_in_data, output_file)
        
        return True
    
    except Exception as e:
        print(f"Error processing device prices for {device_title}: {e}")
        import traceback
        traceback.print_exc()
        return False

def process_brand(driver, wait, brand_name, brand_url, processed_devices, output_file, n_scrape=None, total_scrape_count=0):
    """Process all devices for a specific brand."""
    print(f"\n========== Processing {brand_name} ==========\n")
    
    driver.get(brand_url)
    print(f"Loaded brand page for {brand_name}")
    
    time.sleep(5)
    
    page_num = 1
    continue_processing = True
    
    while continue_processing:
        print(f"Processing page {page_num} for {brand_name}")
        
        try:
            device_cards = WebDriverWait(driver, 10).until(
                EC.presence_of_all_elements_located((By.CSS_SELECTOR, "div[data-v-b4522aa0][data-v-545eeb04]"))
            )
            
            print(f"Found {len(device_cards)} device cards on page {page_num}")
            
            card_index = 0
            while card_index < len(device_cards):
                try:
                    # Re-fetch device cards to avoid stale references
                    device_cards = WebDriverWait(driver, 10).until(
                        EC.presence_of_all_elements_located((By.CSS_SELECTOR, "div[data-v-b4522aa0][data-v-545eeb04]"))
                    )
                    
                    card = device_cards[card_index]
                    
                    title_element = card.find_element(By.CSS_SELECTOR, "span.my-3.text-secondary")
                    title = title_element.text.strip()
                    
                    print(f"\n=== Processing: {title} ===")
                    
                    device_key = f"{brand_name}-{title}"
                    
                    if device_key in processed_devices:
                        print(f"Already processed {title}, skipping")
                        card_index += 1
                        continue
                    
                    sell_button = card.find_element(By.CSS_SELECTOR, "button.btn")
                    
                    driver.execute_script("arguments[0].scrollIntoView({block: 'center'});", sell_button)
                    time.sleep(0.5)
                    
                    print(f"Clicking on {title}")
                    sell_button.click()
                    
                    # Wait for the device page to load
                    wait.until(EC.presence_of_element_located((By.XPATH, "//div[contains(@class, 'box-shadow rounded mb-5')]")))
                    time.sleep(2)
                    
                    # Process the device and get prices
                    process_success = process_device_with_prices(driver, wait, brand_name, title, output_file)
                    
                    processed_devices.append(device_key)
                    total_scrape_count += 1
                    
                    # Return to brand page
                    driver.get(brand_url)
                    print(f"Returned to brand page")
                    time.sleep(5)
                    
                    if page_num > 1:
                        for _ in range(page_num - 1):
                            if not go_to_next_page(driver, wait):
                                print(f"Failed to return to page {page_num}")
                                break
                            time.sleep(1)
                    
                    if n_scrape is not None and total_scrape_count >= n_scrape:
                        print(f"Completed {n_scrape} scrapes as requested")
                        return total_scrape_count
                    
                    # Re-fetch device cards after returning to page
                    device_cards = WebDriverWait(driver, 10).until(
                        EC.presence_of_all_elements_located((By.CSS_SELECTOR, "div[data-v-b4522aa0][data-v-545eeb04]"))
                    )
                    
                    card_index += 1
                
                except StaleElementReferenceException:
                    print(f"Stale element encountered at index {card_index}, refreshing elements")
                    time.sleep(1)
                    # Re-fetch device cards
                    device_cards = WebDriverWait(driver, 10).until(
                        EC.presence_of_all_elements_located((By.CSS_SELECTOR, "div[data-v-b4522aa0][data-v-545eeb04]"))
                    )
                    continue
                
                except Exception as e:
                    print(f"Error processing {title if 'title' in locals() else 'device'}: {e}")
                    driver.get(brand_url)
                    print(f"Returned to brand page due to error")
                    time.sleep(5)
                    
                    if page_num > 1:
                        for _ in range(page_num - 1):
                            if not go_to_next_page(driver, wait):
                                print(f"Failed to return to page {page_num}")
                                break
                            time.sleep(1)
                    
                    # Re-fetch device cards
                    device_cards = WebDriverWait(driver, 10).until(
                        EC.presence_of_all_elements_located((By.CSS_SELECTOR, "div[data-v-b4522aa0][data-v-545eeb04]"))
                    )
                    continue
            
            if go_to_next_page(driver, wait):
                page_num += 1
            else:
                print(f"No more pages for {brand_name}")
                continue_processing = False
                
        except Exception as e:
            print(f"Error processing page {page_num} for {brand_name}: {e}")
            continue_processing = False
    
    return total_scrape_count

def main_loop(n_scrape=None, output_file=None):
    """Main loop to iterate through brands and smartphone models."""
    brand_urls = {
        "Apple": "https://www.trade-mobile.com/home?brand=Apple&search=",
        "Samsung": "https://www.trade-mobile.com/home?brand=Samsung&search="
    }
    
    if output_file is None:
        output_dir = os.environ.get("OUTPUT_DIR", "output")
        os.makedirs(output_dir, exist_ok=True)
        output_file = os.path.join(output_dir, "TH_RV_Source4.xlsx")
    
    print(f"Will save results to: {output_file}")

    driver = setup_driver(headless=True)
    wait = WebDriverWait(driver, 15)

    total_scrape_count = 0
    processed_devices = []
    
    try:
        for brand_name, brand_url in brand_urls.items():
            total_scrape_count = process_brand(
                driver, wait, brand_name, brand_url, 
                processed_devices, output_file, 
                n_scrape, total_scrape_count
            )
            
            if n_scrape is not None and total_scrape_count >= n_scrape:
                break

    except Exception as e:
        print(f"Error in main loop: {e}")
        import traceback
        traceback.print_exc()
        error_folder = os.path.dirname(output_file)
        os.makedirs(error_folder, exist_ok=True)
        driver.save_screenshot(os.path.join(error_folder, "main_loop_error.png"))
    finally:
        driver.quit()
        print(f"Browser closed. Process complete. Processed {total_scrape_count} devices.")

if __name__ == "__main__":
    parser = argparse.ArgumentParser(description='Scrape trade-in values for devices from Thailand Trade-Mobile website')
    parser.add_argument('-n', type=int, help='Number of scrapes to perform (e.g., -n 2 will scrape 2 devices)', default=None)
    parser.add_argument('-o', '--output', type=str, help='Output Excel file path', default=None)
    args = parser.parse_args()
    
    main_loop(args.n, args.output)