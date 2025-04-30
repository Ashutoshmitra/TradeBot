from selenium import webdriver
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.common.exceptions import TimeoutException, NoSuchElementException, StaleElementReferenceException
from webdriver_manager.chrome import ChromeDriverManager
import time
import openpyxl
from datetime import datetime
import os
import re
import argparse

def setup_driver(headless=True):
    """Set up the Chrome WebDriver with appropriate options for containerized environment."""
    options = webdriver.ChromeOptions()
    
    # Only enable headless mode if specified
    if headless:
        options.add_argument('--headless')
    
    # Core options for stability
    options.add_argument('--no-sandbox')
    options.add_argument('--disable-dev-shm-usage')
    options.add_argument('--disable-gpu')
    options.add_argument('--start-maximized')
    options.add_argument('--window-size=1920,1080')
    
    # Performance options
    options.add_argument('--disable-extensions')
    options.add_argument('--disable-infobars')
    options.add_argument('--disable-logging')
    options.add_argument('--disable-notifications')
    options.add_argument('--enable-javascript')
    
    # Cache settings (from successful scripts)
    options.add_argument('--disk-cache-size=1048576')
    options.add_argument('--media-cache-size=1048576')
    
    # User agent
    options.add_argument('--user-agent=Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/91.0.4472.124 Safari/537.36')
    
    # Use a normal page load strategy instead of eager
    options.page_load_strategy = 'normal'
    
    # Initialize the driver directly
    driver = webdriver.Chrome(options=options)
    
    # Set page load timeout to be more generous
    driver.set_page_load_timeout(60)
    
    return driver

def get_dropdown_options(driver, input_id):
    """Retrieve all available options from a dropdown."""
    try:
        dropdown = WebDriverWait(driver, 10).until(
            EC.element_to_be_clickable((By.ID, input_id))
        )
        driver.execute_script("arguments[0].scrollIntoView({block: 'center'});", dropdown)
        dropdown.click()
        time.sleep(1)
        options = driver.find_elements(By.CSS_SELECTOR, f"div[id^='{input_id.replace('input', 'option')}-']")
        option_texts = [option.text.strip() for option in options if option.text.strip()]
        
        # Close dropdown by clicking again
        dropdown.click()
        time.sleep(0.5)
        return option_texts
    except Exception as e:
        print(f"Error getting options for {input_id}: {e}")
        return []

def select_dropdown_option(driver, input_id, option_index, wait, trade_in_data=None):
    """Select an option from a dropdown by index using multiple methods."""
    try:
        print(f"Selecting option {option_index} from dropdown {input_id}")
        
        # Method 1: Standard Selenium approach
        dropdown = wait.until(EC.element_to_be_clickable((By.ID, input_id)))
        driver.execute_script("arguments[0].scrollIntoView({block: 'center'});", dropdown)
        dropdown.click()
        time.sleep(1)
        
        option_selector = f"//div[@id='{input_id.replace('input', 'option')}-{option_index}']"
        option = wait.until(EC.element_to_be_clickable((By.XPATH, option_selector)))
        option_text = option.text.strip()
        
        # Update trade-in data if provided
        if trade_in_data is not None:
            if "react-select-2-input" in input_id:
                trade_in_data["Brand"] = option_text
            elif "react-select-3-input" in input_id:
                trade_in_data["Model"] = option_text
            elif "react-select-4-input" in input_id:
                # For variant information, extract capacity if possible
                variant_text = option_text
                
                # Extract capacity if available (like 128GB or 256GB)
                capacity_match = re.search(r'\d+\s*(?:GB|TB)(?=/|$)|(?<=/)\d+\s*(?:GB|TB)', variant_text)
                if capacity_match:
                    all_matches = re.findall(r'\d+\s*(?:GB|TB)', variant_text)
                    trade_in_data["Capacity"] = all_matches[-1] if all_matches else ""
                else:
                    trade_in_data["Capacity"] = ""
                
        option.click()
        print(f"Selected option: {option_text}")
        return True
    except Exception as e:
        print(f"Standard selection failed: {e}")
        
        # Method 2: JavaScript approach
        try:
            print("Trying JavaScript approach...")
            driver.execute_script(f"""
                var dropdown = document.getElementById('{input_id}');
                if (dropdown) {{
                    dropdown.scrollIntoView({{block: 'center'}});
                    dropdown.click();
                    setTimeout(() => {{
                        var option = document.querySelector('div[id="{input_id.replace("input", "option")}-{option_index}"]');
                        if (option) {{
                            option.click();
                        }}
                    }}, 1000);
                }}
            """)
            time.sleep(2)
            return True
        except Exception as js_error:
            print(f"JavaScript approach failed: {js_error}")
            return False

def navigate_and_complete_form(driver, wait, brand_index, model_index, variant_index, screen_condition, device_type, output_file):
    """Navigate the website and complete the form for a specific configuration."""
    screen_condition_id = f"LCDS-01-{screen_condition}"
    print(f"Processing: {device_type} - Brand index {brand_index}, Model index {model_index}, Variant index {variant_index}, Screen condition: {screen_condition}")

    # Initialize data dictionary with the required columns
    trade_in_data = {
        "Country": "Singapore",
        "Device Type": device_type,
        "Brand": "",
        "Model": "",
        "Capacity": "",
        "Color": "",  # Left blank as requested
        "Launch RRP": "",  # Left blank as requested
        "Condition": "Good" if screen_condition == "minor_scratches" else
             "Damaged" if screen_condition == "cracked" else
             screen_condition.replace("_", " ").title(),
        "Value Type": "Trade-in",
        "Currency": "SGD",
        "Value": "",
        "Source": "SG_RV_Source3",
        "Updated on": datetime.now().strftime("%Y-%m-%d"),
        "Updated by": "",  # Left blank as requested
        "Comments": ""  # Left blank as requested
    }

    try:
        driver.get("https://starhubtradein-sg.compasia.com/")
        
        # FIXED: Click the appropriate device type button with better selector
        device_button_text = "Smartphone" if device_type in ["Phone", "Smartphone", "SmartPhone"] else device_type
        
        # Try multiple methods to click the device type button
        try:
            # Method 1: Try clicking on the parent card button
            wait.until(EC.element_to_be_clickable(
                (By.XPATH, f"//div[contains(@class, 'card-button-footer') and contains(text(), '{device_button_text}')]/parent::div")
            )).click()
            print(f"Successfully clicked on {device_button_text} using parent card selector")
        except Exception as e:
            print(f"First method failed: {e}")
            try:
                # Method 2: Try JavaScript approach to click the button
                script = f"""
                    var buttons = document.querySelectorAll('.card-button');
                    for (var i = 0; i < buttons.length; i++) {{
                        var footer = buttons[i].querySelector('.card-button-footer');
                        if (footer && footer.textContent.includes('{device_button_text}')) {{
                            buttons[i].click();
                            return true;
                        }}
                    }}
                    return false;
                """
                clicked = driver.execute_script(script)
                if clicked:
                    print(f"Successfully clicked on {device_button_text} using JavaScript")
                else:
                    print(f"Could not find {device_button_text} button with JavaScript")
            except Exception as js_error:
                print(f"JavaScript approach failed: {js_error}")
        
        time.sleep(2)

        # Select brand
        select_dropdown_option(driver, "react-select-2-input", brand_index, wait, trade_in_data)
        time.sleep(2)
        
        # Select model
        select_dropdown_option(driver, "react-select-3-input", model_index, wait, trade_in_data)
        time.sleep(2)
        
        # Select variant
        select_dropdown_option(driver, "react-select-4-input", variant_index, wait, trade_in_data)
        time.sleep(2)

        # Click next button
        try:
            next_button = wait.until(EC.element_to_be_clickable(
                (By.XPATH, "//button[contains(@class, 'progress-button-next') and not(@disabled)]")
            ))
            driver.execute_script("arguments[0].scrollIntoView({block: 'center'});", next_button)
            next_button.click()
            print("Successfully clicked Next button")
        except Exception as e:
            print(f"Failed to click Next button: {e}")
            driver.execute_script("""
                var nextBtn = document.querySelector('button.progress-button-next:not([disabled])');
                if (nextBtn) {
                    nextBtn.scrollIntoView({block: 'center'});
                    nextBtn.click();
                }
            """)
            time.sleep(1)

        time.sleep(2)
        
        # Click Skip button on the third page
        time.sleep(3)  # Add longer pause to make sure page loads completely
        
        try:
            # Use direct XPATH to the Skip button based on exact text and class
            skip_button = wait.until(EC.element_to_be_clickable(
                (By.XPATH, "//button[contains(@class, 'progress-button-next') and text()='Skip']")
            ))
            driver.execute_script("arguments[0].scrollIntoView({block: 'center'});", skip_button)
            time.sleep(1)  # Add a small pause before clicking
            driver.execute_script("arguments[0].click();", skip_button)  # Click using JavaScript
            print("Successfully clicked Skip button using direct XPATH")
            time.sleep(2)  # Give more time after clicking
        except Exception as e:
            print(f"Failed to click Skip button: {e}")
            # Take a screenshot for debugging if the skip button click fails
            driver.save_screenshot(f"skip_button_error_{device_type}_{brand_index}_{model_index}_{variant_index}.png")

        time.sleep(2)

        # Fill diagnostic form using JavaScript
        driver.execute_script(f"""
            function clickYesButton(labelText) {{
                var labels = document.querySelectorAll('label.diagnostic-form-label');
                for (var i = 0; i < labels.length; i++) {{
                    if (labels[i].textContent.includes(labelText)) {{
                        var yesButton = labels[i].parentNode.querySelector('button:first-of-type');
                        if (yesButton) {{
                            yesButton.scrollIntoView({{block: 'center'}});
                            yesButton.click();
                        }}
                        break;
                    }}
                }}
            }}
            
            // Device locks
            clickYesButton('Is your device free of any locks');
            
            // Screen condition
            var screenLabel = document.querySelector('label[for="{screen_condition_id}"]');
            if (screenLabel) {{
                screenLabel.scrollIntoView({{block: 'center'}});
                screenLabel.click();
            }}
            
            // Body condition
            var bodyLabel = document.querySelector('label[for="DECO-01-flawless"]');
            if (bodyLabel) {{
                bodyLabel.scrollIntoView({{block: 'center'}});
                bodyLabel.click();
            }}
            
            // Other conditions
            clickYesButton('Fingerprint/Face ID working');
            clickYesButton('device functions below working fine');
            clickYesButton('front and back cameras');
            
            // None of the above checkbox
            var labels = document.querySelectorAll('label');
            for (var i = 0; i < labels.length; i++) {{
                if (labels[i].textContent.trim().includes('None of the above')) {{
                    var checkbox = labels[i].previousElementSibling;
                    if (!checkbox || checkbox.type !== 'checkbox') {{
                        var parent = labels[i].parentElement;
                        checkbox = parent.querySelector('input[type="checkbox"]');
                    }}
                    if (checkbox && checkbox.type === 'checkbox') {{
                        checkbox.scrollIntoView({{block: 'center'}});
                        checkbox.checked = true;
                        checkbox.click();
                        checkbox.dispatchEvent(new Event('change', {{ bubbles: true }}));
                    }}
                    break;
                }}
            }}
        """)
        time.sleep(2)

        # Click Get Quote button
        try:
            quote_button = wait.until(EC.element_to_be_clickable(
                (By.XPATH, "//button[@type='submit' and contains(text(), 'Get Quote')]")
            ))
            driver.execute_script("arguments[0].scrollIntoView({block: 'center'});", quote_button)
            time.sleep(0.5)
            quote_button.click()
            print("Clicked Get Quote button")
        except Exception as e:
            print(f"Failed to click Get Quote button: {e}")
            driver.execute_script("""
                var quoteBtn = document.querySelector('button[type="submit"]');
                if (quoteBtn) {
                    quoteBtn.scrollIntoView({block: 'center'});
                    quoteBtn.click();
                }
            """)
            time.sleep(1)

        time.sleep(5)

        # Extract trade-in value
        try:
            wait.until(EC.visibility_of_element_located((By.CLASS_NAME, "pricing-display-table")))
            
            # Currency is set to SGD by default, but extract it if available
            currency_element = driver.find_element(By.CLASS_NAME, "price-product-name.currency")
            if currency_element:
                currency = currency_element.text.strip()
                if currency:
                    trade_in_data["Currency"] = currency

            # Extract the price
            price_element = driver.find_element(By.CLASS_NAME, "pricing-display-price")
            price_text = price_element.text.strip() if price_element else ""
            price_clean = re.sub(r'[^0-9.]', '', price_text)
            trade_in_data["Value"] = price_clean

            print(f"Extracted trade-in value: {trade_in_data['Currency']} {price_clean}")
            save_to_excel(trade_in_data, output_file)
            
            return True

        except Exception as e:
            print(f"Failed to extract trade-in value: {e}")
            driver.save_screenshot(f"error_{device_type}_{brand_index}_{model_index}_{variant_index}_{screen_condition}.png")
            return False

    except Exception as main_error:
        print(f"Main execution failed: {main_error}")
        driver.save_screenshot(f"error_{device_type}_{brand_index}_{model_index}_{variant_index}_{screen_condition}.png")
        return False

def save_to_excel(data, output_file):
    """Save the extracted trade-in data to an Excel file."""
    # Create output directory if it doesn't exist
    os.makedirs(os.path.dirname(output_file) if os.path.dirname(output_file) else ".", exist_ok=True)
    
    if os.path.exists(output_file):
        try:
            workbook = openpyxl.load_workbook(output_file)
            sheet = workbook.active
        except:
            workbook = openpyxl.Workbook()
            sheet = workbook.active
            # Use the new column headers
            headers = ["Country", "Device Type", "Brand", "Model", "Capacity", "Color", 
                      "Launch RRP", "Condition", "Value Type", "Currency", "Value", 
                      "Source", "Updated on", "Updated by", "Comments"]
            sheet.append(headers)
    else:
        workbook = openpyxl.Workbook()
        sheet = workbook.active
        # Use the new column headers
        headers = ["Country", "Device Type", "Brand", "Model", "Capacity", "Color", 
                  "Launch RRP", "Condition", "Value Type", "Currency", "Value", 
                  "Source", "Updated on", "Updated by", "Comments"]
        sheet.append(headers)

    # Add new row with all the required columns
    row_data = [
        data["Country"],
        data["Device Type"],
        data["Brand"],
        data["Model"],
        data["Capacity"],
        data["Color"],
        data["Launch RRP"],
        data["Condition"],
        data["Value Type"],
        data["Currency"],
        data["Value"],
        data["Source"],
        data["Updated on"],
        data["Updated by"],
        data["Comments"]
    ]
    sheet.append(row_data)

    try:
        workbook.save(output_file)
        print(f"Saved data to {output_file}")
    except Exception as e:
        print(f"Error saving Excel file: {e}")
        alt_file_name = f"SG_RV_Source3.xlsx"
        try:
            workbook.save(alt_file_name)
            print(f"Saved to {alt_file_name}")
        except:
            print("Could not save Excel file")

def process_device_type(driver, wait, device_type, brands, screen_conditions, n_scrape=None, output_file=None):
    """Process a specific device type (Phone or Tablet)."""
    device_button_text = "Smartphone" if device_type in ["Phone", "Smartphone", "SmartPhone"] else device_type
    print(f"\n=== Starting {device_type} scraping ===\n")
    
    # Counter for number of scrapes completed for this device type
    scrape_count = 0
    
    try:
        # FIXED: Better device selection approach
        driver.get("https://starhubtradein-sg.compasia.com/")
        
        # Try multiple methods to click the device type button
        try:
            # Method 1: Try clicking on the parent card button
            wait.until(EC.element_to_be_clickable(
                (By.XPATH, f"//div[contains(@class, 'card-button-footer') and contains(text(), '{device_button_text}')]/parent::div")
            )).click()
            print(f"Successfully clicked on {device_button_text} using parent card selector")
        except Exception as e:
            print(f"First method failed: {e}")
            try:
                # Method 2: Try JavaScript approach to click the button
                script = f"""
                    var buttons = document.querySelectorAll('.card-button');
                    for (var i = 0; i < buttons.length; i++) {{
                        var footer = buttons[i].querySelector('.card-button-footer');
                        if (footer && footer.textContent.includes('{device_button_text}')) {{
                            buttons[i].click();
                            return true;
                        }}
                    }}
                    return false;
                """
                clicked = driver.execute_script(script)
                if clicked:
                    print(f"Successfully clicked on {device_button_text} using JavaScript")
                else:
                    print(f"Could not find {device_button_text} button with JavaScript")
            except Exception as js_error:
                print(f"JavaScript approach failed: {js_error}")
        
        time.sleep(2)

        # Get all brand options and find indices for the brands we're interested in
        brand_options = get_dropdown_options(driver, "react-select-2-input")
        print(f"All available brands for {device_type}: {brand_options}")
        
        brand_indices = []
        for brand in brands:
            try:
                idx = next((i for i, b in enumerate(brand_options) if b == brand), None)
                if idx is not None:
                    brand_indices.append(idx)
                    print(f"Found {brand} at index {idx}")
                else:
                    print(f"Brand {brand} not found in options for {device_type}")
            except Exception as e:
                print(f"Error finding brand {brand} for {device_type}: {e}")
                continue

        # Main loop for all combinations for this device type
        for brand_idx in brand_indices:
            # Navigate to the main page for each brand
            driver.get("https://starhubtradein-sg.compasia.com/")
            
            # FIXED: Better device selection with multiple approaches
            try:
                # Method 1: Try clicking on the parent card button
                wait.until(EC.element_to_be_clickable(
                    (By.XPATH, f"//div[contains(@class, 'card-button-footer') and contains(text(), '{device_button_text}')]/parent::div")
                )).click()
                print(f"Successfully clicked on {device_button_text} using parent card selector")
            except Exception as e:
                print(f"First method failed: {e}")
                try:
                    # Method 2: Try JavaScript approach to click the button
                    script = f"""
                        var buttons = document.querySelectorAll('.card-button');
                        for (var i = 0; i < buttons.length; i++) {{
                            var footer = buttons[i].querySelector('.card-button-footer');
                            if (footer && footer.textContent.includes('{device_button_text}')) {{
                                buttons[i].click();
                                return true;
                            }}
                        }}
                        return false;
                    """
                    clicked = driver.execute_script(script)
                    if clicked:
                        print(f"Successfully clicked on {device_button_text} using JavaScript")
                    else:
                        print(f"Could not find {device_button_text} button with JavaScript")
                except Exception as js_error:
                    print(f"JavaScript approach failed: {js_error}")
            
            time.sleep(2)
            
            # Select the brand
            select_dropdown_option(driver, "react-select-2-input", brand_idx, wait)
            time.sleep(2)
            
            # Get available models for this brand
            model_options = get_dropdown_options(driver, "react-select-3-input")
            num_models = len(model_options)
            print(f"Found {num_models} models for {device_type} brand index {brand_idx}: {model_options}")
            
            for model_idx in range(num_models):
                # For each model, go back to the main page
                driver.get("https://starhubtradein-sg.compasia.com/")
                
                # FIXED: Better device selection with multiple approaches
                try:
                    # Method 1: Try clicking on the parent card button
                    wait.until(EC.element_to_be_clickable(
                        (By.XPATH, f"//div[contains(@class, 'card-button-footer') and contains(text(), '{device_button_text}')]/parent::div")
                    )).click()
                    print(f"Successfully clicked on {device_button_text} using parent card selector")
                except Exception as e:
                    print(f"First method failed: {e}")
                    try:
                        # Method 2: Try JavaScript approach to click the button
                        script = f"""
                            var buttons = document.querySelectorAll('.card-button');
                            for (var i = 0; i < buttons.length; i++) {{
                                var footer = buttons[i].querySelector('.card-button-footer');
                                if (footer && footer.textContent.includes('{device_button_text}')) {{
                                    buttons[i].click();
                                    return true;
                                }}
                            }}
                            return false;
                        """
                        clicked = driver.execute_script(script)
                        if clicked:
                            print(f"Successfully clicked on {device_button_text} using JavaScript")
                        else:
                            print(f"Could not find {device_button_text} button with JavaScript")
                    except Exception as js_error:
                        print(f"JavaScript approach failed: {js_error}")
                
                time.sleep(2)
                
                # Select brand again
                select_dropdown_option(driver, "react-select-2-input", brand_idx, wait)
                time.sleep(2)
                
                # Select model
                select_dropdown_option(driver, "react-select-3-input", model_idx, wait)
                time.sleep(2)
                
                # Get variants for this model
                variant_options = get_dropdown_options(driver, "react-select-4-input")
                num_variants = len(variant_options)
                print(f"Found {num_variants} variants for {device_type} model index {model_idx}: {variant_options}")
                
                for variant_idx in range(num_variants):
                    for condition in screen_conditions:
                        print(f"\nStarting new configuration: {device_type} - Brand idx {brand_idx}, Model idx {model_idx}, Variant idx {variant_idx}, Condition: {condition}")
                        result = navigate_and_complete_form(driver, wait, brand_idx, model_idx, variant_idx, condition, device_type, output_file)
                        time.sleep(2)  # Brief pause between iterations
                        
                        # Increment the scrape counter
                        scrape_count += 1
                        
                        # If there was an error, retry once
                        if not result:
                            print(f"Retrying {device_type} after error...")
                            navigate_and_complete_form(driver, wait, brand_idx, model_idx, variant_idx, condition, device_type, output_file)
                            time.sleep(2)
                            
                        # Check if we've reached the requested number of scrapes for this device type
                        if n_scrape is not None and scrape_count >= n_scrape:
                            print(f"Completed {scrape_count} {device_type} scrapes as requested. Stopping.")
                            return scrape_count
    
    except Exception as e:
        print(f"Error in {device_type} processing: {e}")
        driver.save_screenshot(f"{device_type}_processing_error.png")
    
    return scrape_count

def main_loop(n_scrape=None, output_file=None):
    """Main loop to iterate through device types, brands, models, variants, and screen conditions."""
    # Common brands to check for both phones and tablets
    # brands = ["Apple", "Samsung", "Google", "Huawei", "Xiaomi", "Oppo", "OnePlus", "Sony", "LG", "Motorola", "Vivo", "Realme", "Honor", "Nubia", "Nothing"]
    brands = ["Apple", "Samsung"]
    screen_conditions = ["flawless", "minor_scratches", "cracked"]

    # Use default output path if not specified
    if output_file is None:
        # Check for environment variable first
        output_dir = os.environ.get("OUTPUT_DIR", "output")
        os.makedirs(output_dir, exist_ok=True)
        output_file = os.path.join(output_dir, "SG_RV_Source3.xlsx")

    # Create output directory if it doesn't exist
    os.makedirs(os.path.dirname(output_file) if os.path.dirname(output_file) else ".", exist_ok=True)
    print(f"Will save results to: {output_file}")

    # Setup driver (single browser instance)
    driver = setup_driver(headless=True)
    
    ignored_exceptions = (NoSuchElementException, StaleElementReferenceException)
    wait = WebDriverWait(driver, 15, 0.5, ignored_exceptions=ignored_exceptions)

    try:
        # If n_scrape is specified, split it between device types
        phones_to_scrape = tablets_to_scrape = None
        if n_scrape is not None:
            phones_to_scrape = n_scrape // 2  # Half for phones
            tablets_to_scrape = n_scrape - phones_to_scrape  # Remaining for tablets
            print(f"Planning to scrape {phones_to_scrape} phones and {tablets_to_scrape} tablets")
        
        # Process smartphones
        phones_scraped = process_device_type(driver, wait, "SmartPhone", brands, screen_conditions, phones_to_scrape, output_file)
        
        # Process tablets
        tablets_scraped = process_device_type(driver, wait, "Tablet", brands, screen_conditions, tablets_to_scrape, output_file)
        
        # Report total
        total_scraped = phones_scraped + tablets_scraped
        print(f"\n=== Scraping completed ===")
        print(f"Total items scraped: {total_scraped} ({phones_scraped} phones, {tablets_scraped} tablets)")

    except Exception as e:
        print(f"Error in main loop: {e}")
        driver.save_screenshot("main_loop_error.png")
    finally:
        driver.quit()
        print("Browser closed. Process complete.")

if __name__ == "__main__":
    parser = argparse.ArgumentParser(description='Scrape StarHub trade-in values for smartphones and tablets')
    parser.add_argument('-n', type=int, help='Number of scrapes to perform (divided between phones and tablets)', default=None)
    parser.add_argument('-o', '--output', type=str, help='Output Excel file path', default=None)
    args = parser.parse_args()
    
    main_loop(args.n, args.output)