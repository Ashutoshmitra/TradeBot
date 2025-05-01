from selenium import webdriver
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.common.exceptions import TimeoutException, NoSuchElementException, StaleElementReferenceException
from webdriver_manager.chrome import ChromeDriverManager
import time
import openpyxl
from datetime import datetime
import os
import re
import argparse

def setup_driver(headless=True):
    """Set up the Chrome WebDriver with appropriate options for containerized environment."""
    options = webdriver.ChromeOptions()
    
    # Only enable headless mode if specified
    if headless:
        options.add_argument('--headless')
    
    # Core options for stability
    options.add_argument('--no-sandbox')
    options.add_argument('--disable-dev-shm-usage')
    options.add_argument('--disable-gpu')
    options.add_argument('--start-maximized')
    options.add_argument('--window-size=1920,1080')
    
    # Performance options
    options.add_argument('--disable-extensions')
    options.add_argument('--disable-infobars')
    options.add_argument('--disable-logging')
    options.add_argument('--disable-notifications')
    options.add_argument('--enable-javascript')
    
    # Cache settings (from successful scripts)
    options.add_argument('--disk-cache-size=1048576')
    options.add_argument('--media-cache-size=1048576')
    
    # User agent
    options.add_argument('--user-agent=Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/91.0.4472.124 Safari/537.36')
    
    # Use a normal page load strategy instead of eager
    options.page_load_strategy = 'normal'
    
    # Initialize the driver directly
    driver = webdriver.Chrome(options=options)
    
    # Set page load timeout to be more generous
    driver.set_page_load_timeout(60)
    
    return driver


def get_condition_mapping(screen_condition):
    """Map screen condition values to the required conditions."""
    condition_map = {
        "flawless": "Flawless",
        "minor_scratches": "Good",
        "cracked": "Damaged"
    }
    return condition_map.get(screen_condition, screen_condition.replace("_", " ").title())


def click_device_type(driver, wait, device_type):
    """Click on the device type card with multiple fallback methods for Malaysia website."""
    print(f"Attempting to click on {device_type} card...")
    
    # Check for iframes first
    print("Checking for iframes...")
    iframes = driver.find_elements(By.TAG_NAME, 'iframe')
    found_in_iframe = False
    
    if iframes:
        print(f"Found {len(iframes)} iframes")
        for index, iframe in enumerate(iframes):
            try:
                driver.switch_to.frame(iframe)
                print(f"Switched to iframe {index}")
                # Check if cards exist in this iframe
                cards = WebDriverWait(driver, 5).until(
                    EC.presence_of_all_elements_located((By.CSS_SELECTOR, '.card-button'))
                )
                print(f"Found cards in iframe {index}")
                found_in_iframe = True
                break
            except TimeoutException:
                print(f"No cards found in iframe {index}")
                driver.switch_to.default_content()
    
    if not found_in_iframe:
        driver.switch_to.default_content()
        print("No cards found in iframes, proceeding with main content")
    
    # Method 1: Find by card-button-footer text
    try:
        cards = WebDriverWait(driver, 15).until(
            EC.visibility_of_all_elements_located((By.CSS_SELECTOR, '.card-button'))
        )
        print(f"Found {len(cards)} cards")
        
        card = None
        for c in cards:
            try:
                footer = c.find_element(By.CSS_SELECTOR, '.card-button-footer')
                if footer.text.strip() == device_type:
                    card = c
                    break
            except (StaleElementReferenceException, NoSuchElementException) as e:
                print(f"Error finding footer text: {e}")
                continue
        
        if card:
            # Check if card is disabled
            aria_disabled = card.get_attribute('aria-disabled')
            print(f"aria-disabled for {device_type}: {aria_disabled}")
            
            # Scroll to the card
            driver.execute_script("arguments[0].scrollIntoView({block: 'center'});", card)
            time.sleep(1)
            
            # Try multiple click methods
            try:
                card.click()
                print(f"Clicked {device_type} using standard click")
            except (ElementClickInterceptedException, StaleElementReferenceException):
                print(f"Standard click failed for {device_type}, trying ActionChains")
                ActionChains(driver).move_to_element(card).click().perform()
                print(f"Clicked {device_type} using ActionChains")
            except Exception as e:
                print(f"ActionChains click failed: {e}, trying JavaScript")
                driver.execute_script("arguments[0].click();", card)
                print(f"Clicked {device_type} using JavaScript")
            
            time.sleep(3)  # Wait for page to update
            return True
    except Exception as e:
        print(f"Method 1 failed: {e}")
    
    # Method 2: JavaScript approach
    try:
        script = f"""
            var cards = document.querySelectorAll('.card-button');
            var clicked = false;
            for (var i = 0; i < cards.length; i++) {{
                var footer = cards[i].querySelector('.card-button-footer');
                if (footer && footer.textContent.trim() === '{device_type}') {{
                    cards[i].scrollIntoView({{block: 'center'}});
                    setTimeout(function() {{
                        cards[i].click();
                    }}, 500);
                    clicked = true;
                    break;
                }}
            }}
            return clicked;
        """
        result = driver.execute_script(script)
        if result:
            print(f"Successfully clicked on {device_type} card using JavaScript")
            time.sleep(3)  # Wait for page to update
            return True
    except Exception as e:
        print(f"JavaScript approach failed: {e}")
    
    # Method 3: XPath with parent
    try:
        xpath = f"//div[contains(@class, 'card-button-footer') and text()='{device_type}']/parent::div"
        element = wait.until(EC.element_to_be_clickable((By.XPATH, xpath)))
        driver.execute_script("arguments[0].scrollIntoView({block: 'center'});", element)
        time.sleep(1)
        driver.execute_script("arguments[0].click();", element)
        print(f"Successfully clicked on {device_type} using XPath with parent")
        time.sleep(3)
        return True
    except Exception as e:
        print(f"XPath parent click failed: {e}")
    
    # Save screenshot for debugging
    try:
        driver.save_screenshot(f"click_failure_{device_type}.png")
        print(f"Screenshot saved as click_failure_{device_type}.png")
    except Exception as e:
        print(f"Failed to save screenshot: {e}")
    
    print(f"All methods to click {device_type} card failed")
    return False

def get_dropdown_options(driver, input_id):
    """Retrieve all available options from a dropdown."""
    try:
        dropdown = WebDriverWait(driver, 10).until(
            EC.element_to_be_clickable((By.ID, input_id))
        )
        driver.execute_script("arguments[0].scrollIntoView({block: 'center'});", dropdown)
        dropdown.click()
        time.sleep(1)
        options = driver.find_elements(By.CSS_SELECTOR, f"div[id^='{input_id.replace('input', 'option')}-']")
        option_texts = [option.text.strip() for option in options if option.text.strip()]
        
        # Close dropdown by clicking again
        dropdown.click()
        time.sleep(0.5)
        return option_texts
    except Exception as e:
        print(f"Error getting options for {input_id}: {e}")
        return []

def select_dropdown_option(driver, input_id, option_index, wait, trade_in_data=None):
    """Select an option from a dropdown by index using multiple methods."""
    try:
        print(f"Selecting option {option_index} from dropdown {input_id}")
        
        # Method 1: Standard Selenium approach
        dropdown = wait.until(EC.element_to_be_clickable((By.ID, input_id)))
        driver.execute_script("arguments[0].scrollIntoView({block: 'center'});", dropdown)
        dropdown.click()
        time.sleep(1)
        
        option_selector = f"//div[@id='{input_id.replace('input', 'option')}-{option_index}']"
        option = wait.until(EC.element_to_be_clickable((By.XPATH, option_selector)))
        option_text = option.text.strip()
        
        # Update trade-in data if provided
        if trade_in_data is not None:
            if "react-select-2-input" in input_id:
                trade_in_data["Brand"] = option_text
            elif "react-select-3-input" in input_id:
                trade_in_data["Model"] = option_text
            elif "react-select-4-input" in input_id:
                # For variant information, extract capacity if possible
                variant_text = option_text
                
                # Extract capacity if available (like 128GB or 256GB)
                capacity_match = re.search(r'\d+\s*(?:GB|TB)(?=/|$)|(?<=/)\d+\s*(?:GB|TB)', variant_text)
                if capacity_match:
                    all_matches = re.findall(r'\d+\s*(?:GB|TB)', variant_text)
                    trade_in_data["Capacity"] = all_matches[-1] if all_matches else ""
                else:
                    trade_in_data["Capacity"] = ""
                
        option.click()
        print(f"Selected option: {option_text}")
        return True
    except Exception as e:
        print(f"Standard selection failed: {e}")
        
        # Method 2: JavaScript approach
        try:
            print("Trying JavaScript approach...")
            driver.execute_script(f"""
                var dropdown = document.getElementById('{input_id}');
                if (dropdown) {{
                    dropdown.scrollIntoView({{block: 'center'}});
                    dropdown.click();
                    setTimeout(() => {{
                        var option = document.querySelector('div[id="{input_id.replace("input", "option")}-{option_index}"]');
                        if (option) {{
                            option.click();
                        }}
                    }}, 1000);
                }}
            """)
            time.sleep(2)
            return True
        except Exception as js_error:
            print(f"JavaScript approach failed: {js_error}")
            return False


def click_button_with_fallback(driver, wait, primary_selector, backup_js_selector, button_name):
    """Click a button with a fallback method if primary selector fails."""
    try:
        button = wait.until(EC.element_to_be_clickable((By.XPATH, primary_selector)))
        driver.execute_script("arguments[0].scrollIntoView({block: 'center'});", button)
        driver.execute_script("arguments[0].click();", button)
        print(f"Successfully clicked {button_name} button")
        return True
    except Exception as e:
        print(f"Primary method for clicking {button_name} button failed: {e}")
        
        # JavaScript fallback
        try:
            script = backup_js_selector
            clicked = driver.execute_script(script)
            if clicked:
                print(f"Successfully clicked {button_name} button using JavaScript")
                return True
            else:
                print(f"JavaScript could not find {button_name} button")
                return False
        except Exception as js_error:
            print(f"JavaScript fallback for clicking {button_name} button failed: {js_error}")
            return False


def fill_diagnostic_form(driver, screen_condition):
    """Fill the diagnostic form using JavaScript."""
    screen_condition_id = f"LCDS-01-{screen_condition}"
    
    script = f"""
        function clickYesButton(labelText) {{
            var labels = document.querySelectorAll('label.diagnostic-form-label');
            for (var i = 0; i < labels.length; i++) {{
                if (labels[i].textContent.includes(labelText)) {{
                    var yesButton = labels[i].parentNode.querySelector('button:first-of-type');
                    if (yesButton) {{
                        yesButton.scrollIntoView({{block: 'center'}});
                        yesButton.click();
                    }}
                    break;
                }}
            }}
        }}
        
        // Device locks
        clickYesButton('Is your device free of any locks');
        
        // Screen condition
        var screenLabel = document.querySelector('label[for="{screen_condition_id}"]');
        if (screenLabel) {{
            screenLabel.scrollIntoView({{block: 'center'}});
            screenLabel.click();
        }}
        
        // Body condition
        var bodyLabel = document.querySelector('label[for="DECO-01-flawless"]');
        if (bodyLabel) {{
            bodyLabel.scrollIntoView({{block: 'center'}});
            bodyLabel.click();
        }}
        
        // Other conditions
        clickYesButton('Fingerprint/Face ID working');
        clickYesButton('device functions below working fine');
        clickYesButton('front and back cameras');
        
        // None of the above checkbox
        var labels = document.querySelectorAll('label');
        for (var i = 0; i < labels.length; i++) {{
            if (labels[i].textContent.trim().includes('None of the above')) {{
                var checkbox = labels[i].previousElementSibling;
                if (!checkbox || checkbox.type !== 'checkbox') {{
                    var parent = labels[i].parentElement;
                    checkbox = parent.querySelector('input[type="checkbox"]');
                }}
                if (checkbox && checkbox.type === 'checkbox') {{
                    checkbox.scrollIntoView({{block: 'center'}});
                    checkbox.checked = true;
                    checkbox.click();
                    checkbox.dispatchEvent(new Event('change', {{ bubbles: true }}));
                }}
                break;
            }}
        }}
    """
    
    driver.execute_script(script)
    print("Filled diagnostic form")
    time.sleep(2)


def extract_trade_in_value(driver, wait, trade_in_data, output_file):
    """Extract the trade-in value from the results page."""
    try:
        wait.until(EC.visibility_of_element_located((By.CLASS_NAME, "pricing-display-table")))
        
        # Extract currency using JavaScript
        currency_script = """
            var element = document.querySelector('.price-product-name.currency');
            return element ? element.textContent.trim() : 'SGD';
        """
        currency = driver.execute_script(currency_script)
        trade_in_data["Currency"] = currency
        
        # Extract price using JavaScript
        price_script = """
            var element = document.querySelector('.pricing-display-price');
            return element ? element.textContent.trim() : '';
        """
        price_text = driver.execute_script(price_script)
        price_clean = re.sub(r'[^0-9.]', '', price_text)
        trade_in_data["Value"] = price_clean
        
        print(f"Extracted trade-in value: {currency} {price_clean}")
        save_to_excel(trade_in_data, output_file)
        
        return True
    except Exception as e:
        print(f"Failed to extract trade-in value: {e}")
        return False


def navigate_and_complete_form(driver, wait, device_type, brand_index, model_index, variant_index, screen_condition, output_file):
    """Navigate the website and complete the form for a specific configuration."""
    print(f"Processing: {device_type}, Brand index {brand_index}, Model index {model_index}, Variant index {variant_index}, Screen condition: {screen_condition}")

    # Initialize data dictionary with the required columns
    trade_in_data = {
        "Country": "Singapore",
        "Device Type": device_type,
        "Brand": "",
        "Model": "",
        "Capacity": "",
        "Color": "",
        "Launch RRP": "",
        "Condition": get_condition_mapping(screen_condition),
        "Value Type": "Trade-in",
        "Currency": "SGD",
        "Value": "",
        "Source": "SG_RV_Source1",
        "Updated on": datetime.now().strftime("%Y-%m-%d"),
        "Updated by": "",
        "Comments": ""
    }

    try:
        # Load the main page
        driver.get("https://compasia.my/pages/sell-your-devices")
        time.sleep(3)
        
        # Click on device type card
        if not click_device_type(driver, wait, device_type):
            return False
        
        # Select brand
        if not select_dropdown_option(driver, "react-select-2-input", brand_index, wait, trade_in_data):
            return False
        
        # Select model
        if not select_dropdown_option(driver, "react-select-3-input", model_index, wait, trade_in_data):
            return False
        
        # Select variant
        if not select_dropdown_option(driver, "react-select-4-input", variant_index, wait, trade_in_data):
            return False
        
        # Click next button
        next_button_js = """
            var nextBtn = document.querySelector('button.progress-button-next:not([disabled])');
            if (nextBtn) {
                nextBtn.scrollIntoView({block: 'center'});
                nextBtn.click();
                return true;
            }
            return false;
        """
        if not click_button_with_fallback(
            driver, wait, 
            "//button[contains(@class, 'progress-button-next') and not(@disabled)]", 
            next_button_js, "Next"
        ):
            return False
        
        time.sleep(3)
        
        # Fill diagnostic form
        fill_diagnostic_form(driver, screen_condition)
        
        # Click Get Quote button
        quote_button_js = """
            var quoteBtn = document.querySelector('button[type="submit"]');
            if (quoteBtn) {
                quoteBtn.scrollIntoView({block: 'center'});
                quoteBtn.click();
                return true;
            }
            return false;
        """
        if not click_button_with_fallback(
            driver, wait, 
            "//button[@type='submit' and contains(text(), 'Get Quote')]", 
            quote_button_js, "Get Quote"
        ):
            return False
        
        time.sleep(5)
        
        # Extract trade-in value
        return extract_trade_in_value(driver, wait, trade_in_data, output_file)
    
    except Exception as main_error:
        print(f"Main execution failed: {main_error}")
        error_folder = os.path.dirname(output_file)
        os.makedirs(error_folder, exist_ok=True)
        error_screenshot = os.path.join(error_folder, f"error_{device_type}_{brand_index}_{model_index}_{variant_index}_{screen_condition}.png")
        driver.save_screenshot(error_screenshot)
        return False


def save_to_excel(data, output_file):
    """Save the extracted trade-in data to an Excel file."""
    # Ensure output directory exists
    os.makedirs(os.path.dirname(output_file) if os.path.dirname(output_file) else ".", exist_ok=True)
    
    if os.path.exists(output_file):
        try:
            workbook = openpyxl.load_workbook(output_file)
            sheet = workbook.active
        except:
            workbook = openpyxl.Workbook()
            sheet = workbook.active
            # Use the new column headers
            headers = ["Country", "Device Type", "Brand", "Model", "Capacity", "Color", 
                      "Launch RRP", "Condition", "Value Type", "Currency", "Value", 
                      "Source", "Updated on", "Updated by", "Comments"]
            sheet.append(headers)
    else:
        workbook = openpyxl.Workbook()
        sheet = workbook.active
        # Use the new column headers
        headers = ["Country", "Device Type", "Brand", "Model", "Capacity", "Color", 
                  "Launch RRP", "Condition", "Value Type", "Currency", "Value", 
                  "Source", "Updated on", "Updated by", "Comments"]
        sheet.append(headers)

    # Add new row with all the required columns
    row_data = [
        data["Country"],
        data["Device Type"],
        data["Brand"],
        data["Model"],
        data["Capacity"],
        data["Color"],
        data["Launch RRP"],
        data["Condition"],
        data["Value Type"],
        data["Currency"],
        data["Value"],
        data["Source"],
        data["Updated on"],
        data["Updated by"],
        data["Comments"]
    ]
    sheet.append(row_data)

    try:
        workbook.save(output_file)
        print(f"Saved data to {output_file}")
    except Exception as e:
        print(f"Error saving Excel file: {e}")
        alt_file_name = os.path.join(os.path.dirname(output_file), f"SG_RV_Source1.xlsx")
        try:
            workbook.save(alt_file_name)
            print(f"Saved to {alt_file_name}")
        except:
            print("Could not save Excel file")


def main_loop(n_scrape=None, output_file=None):
    """Main loop to iterate through brands, models, variants, and screen conditions."""
    brands = ["Apple", "Samsung"]
    screen_conditions = ["flawless", "minor_scratches", "cracked"]
    device_types = ["Smartphone", "Tablet"]  # Match the exact text from the card-button-footer elements

    # Use default output path if not specified
    if output_file is None:
        # Check for environment variable first
        output_dir = os.environ.get("OUTPUT_DIR", "output")
        os.makedirs(output_dir, exist_ok=True)
        output_file = os.path.join(output_dir, "SG_RV_Source1.xlsx")
    
    print(f"Will save results to: {output_file}")

    # Setup driver (single browser instance)
    driver = setup_driver(headless=True)
    
    ignored_exceptions = (NoSuchElementException, StaleElementReferenceException)
    wait = WebDriverWait(driver, 15, 0.5, ignored_exceptions=ignored_exceptions)

    # Counters for each device type
    device_scrape_counts = {device_type: 0 for device_type in device_types}
    total_scrape_count = 0
    
    try:
        # Loop through each device type (Smartphone and Tablet)
        for device_type in device_types:
            print(f"\n========== Processing {device_type} ==========\n")
            
            # Navigate to website
            driver.get("https://compasia.my/pages/sell-your-devices")
            time.sleep(3)
            
            # Click on the device type
            if not click_device_type(driver, wait, device_type):
                print(f"Could not click on {device_type} card, skipping to next device type")
                continue
            
            # Get all brand options
            brand_options = get_dropdown_options(driver, "react-select-2-input")
            brand_indices = []
            
            for brand in brands:
                try:
                    idx = next((i for i, b in enumerate(brand_options) if b == brand), None)
                    if idx is not None:
                        brand_indices.append(idx)
                        print(f"Found {brand} at index {idx}")
                    else:
                        print(f"Brand {brand} not found in options")
                except Exception as e:
                    print(f"Error finding brand {brand}: {e}")
                    continue
            
            # Process each brand
            for brand_idx in brand_indices:
                # Navigate to the main page for each brand
                driver.get("https://compasia.my/pages/sell-your-devices")
                time.sleep(3)
                
                if not click_device_type(driver, wait, device_type):
                    print(f"Could not click on {device_type} card for brand {brand_idx}, skipping")
                    continue
                
                # Select the brand
                if not select_dropdown_option(driver, "react-select-2-input", brand_idx, wait):
                    print(f"Could not select brand index {brand_idx}, skipping")
                    continue
                
                # Get available models for this brand
                model_options = get_dropdown_options(driver, "react-select-3-input")
                num_models = len(model_options)
                print(f"Found {num_models} models for brand index {brand_idx}: {model_options}")
                
                # Process each model
                for model_idx in range(num_models):
                    # Navigate to main page for each model
                    driver.get("https://compasia.my/pages/sell-your-devices")
                    time.sleep(3)
                    
                    if not click_device_type(driver, wait, device_type):
                        print(f"Could not click on {device_type} card for model {model_idx}, skipping")
                        continue
                    
                    # Select brand again
                    if not select_dropdown_option(driver, "react-select-2-input", brand_idx, wait):
                        print(f"Could not select brand index {brand_idx} for model {model_idx}, skipping")
                        continue
                    
                    # Select model
                    if not select_dropdown_option(driver, "react-select-3-input", model_idx, wait):
                        print(f"Could not select model index {model_idx}, skipping")
                        continue
                    
                    # Get variants for this model
                    variant_options = get_dropdown_options(driver, "react-select-4-input")
                    num_variants = len(variant_options)
                    print(f"Found {num_variants} variants for model index {model_idx}: {variant_options}")
                    
                    # Process each variant and condition
                    for variant_idx in range(num_variants):
                        for condition in screen_conditions:
                            # Check if we've reached the requested number of scrapes for this device type
                            if n_scrape is not None and device_scrape_counts[device_type] >= n_scrape:
                                print(f"Completed {n_scrape} {device_type} scrapes as requested. Moving to next device type.")
                                break  # Break out of conditions loop
                            
                            print(f"\nStarting new configuration: {device_type}, Brand idx {brand_idx}, Model idx {model_idx}, Variant idx {variant_idx}, Condition: {condition}")
                            result = navigate_and_complete_form(driver, wait, device_type, brand_idx, model_idx, variant_idx, condition, output_file)
                            time.sleep(2)  # Brief pause between iterations
                            
                            # If there was an error, retry once
                            if not result:
                                print("Retrying after error...")
                                result = navigate_and_complete_form(driver, wait, device_type, brand_idx, model_idx, variant_idx, condition, output_file)
                                time.sleep(2)
                            
                            # Only increment if successful
                            if result:
                                device_scrape_counts[device_type] += 1
                                total_scrape_count += 1
                                print(f"Completed {device_scrape_counts[device_type]}/{n_scrape if n_scrape else 'unlimited'} {device_type} scrapes")
                        
                        # Check if we should break out of the variant loop
                        if n_scrape is not None and device_scrape_counts[device_type] >= n_scrape:
                            break
                    
                    # Check if we should break out of the model loop
                    if n_scrape is not None and device_scrape_counts[device_type] >= n_scrape:
                        break
                
                # Check if we should break out of the brand loop
                if n_scrape is not None and device_scrape_counts[device_type] >= n_scrape:
                    break
            
            print(f"\nCompleted processing {device_type} - Scraped {device_scrape_counts[device_type]} items")

    except Exception as e:
        print(f"Error in main loop: {e}")
        error_folder = os.path.dirname(output_file)
        os.makedirs(error_folder, exist_ok=True)
        driver.save_screenshot(os.path.join(error_folder, "main_loop_error.png"))
    finally:
        driver.quit()
        print("Browser closed. Process complete.")


if __name__ == "__main__":
    parser = argparse.ArgumentParser(description='Scrape trade-in values for smartphones and tablets')
    parser.add_argument('-n', type=int, help='Number of scrapes to perform per device type (e.g., -n 2 will scrape 2 smartphones and 2 tablets)', default=None)
    parser.add_argument('-o', '--output', type=str, help='Output Excel file path', default=None)
    args = parser.parse_args()
    
    main_loop(args.n, args.output)