from selenium import webdriver
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.common.exceptions import TimeoutException, NoSuchElementException, StaleElementReferenceException
from webdriver_manager.chrome import ChromeDriverManager
import time
import openpyxl
from datetime import datetime
import os
import re
import argparse

def setup_driver():
    """Set up the Chrome WebDriver with appropriate options for containerized environment."""
    options = webdriver.ChromeOptions()
    options.add_argument("--headless")
    options.add_argument("--no-sandbox")
    options.add_argument("--disable-dev-shm-usage")
    options.add_argument("--disable-gpu")
    options.add_argument("--start-maximized")
    options.add_argument("--window-size=1920,1080")
    options.add_argument("--disable-extensions")
    options.add_argument("--disable-popup-blocking")
    options.add_argument("--disable-infobars")
    options.add_argument("--user-agent=Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/91.0.4472.124 Safari/537.36")
    
    # Use Chrome directly instead of webdriver-manager in container
    driver = webdriver.Chrome(options=options)
    return driver

def get_condition_mapping(screen_condition):
    """Map screen condition values to the required conditions."""
    condition_map = {
        "flawless": "Flawless",
        "minor_scratches": "Good",
        "cracked": "Damaged"
    }
    return condition_map.get(screen_condition, screen_condition.replace("_", " ").title())

def click_device_type(driver, wait, device_type):
    """Click on the device type card using a simple, reliable method."""
    print(f"Attempting to click on {device_type} card...")
    
    # Wait a good amount of time for the page to fully load
    time.sleep(5)
    
    # Simple position-based method
    try:
        # Determine the index based on device type
        index = 0
        if device_type.lower() == "smartphone":
            index = 0
        elif device_type.lower() == "tablet":
            index = 1
        elif device_type.lower() == "watch":
            index = 2
            
        # Use JavaScript to click the appropriate card
        script = f"""
            var cards = document.querySelectorAll('.card-button');
            console.log('Found ' + cards.length + ' cards');
            if (cards.length > {index}) {{
                cards[{index}].scrollIntoView({{block: 'center'}});
                setTimeout(function() {{
                    cards[{index}].click();
                }}, 500);
                return true;
            }}
            return false;
        """
        result = driver.execute_script(script)
        if result:
            print(f"Successfully clicked on {device_type} card at position {index}")
            time.sleep(3)  # Wait after clicking
            return True
        else:
            print(f"No card found at position {index}")
    except Exception as e:
        print(f"Position-based method failed: {e}")
    
    # Fallback method - try clicking on any card button
    try:
        print("Trying to click on any card button...")
        script = """
            var cards = document.querySelectorAll('.card-button');
            if (cards.length > 0) {
                cards[0].scrollIntoView({block: 'center'});
                setTimeout(function() {
                    cards[0].click();
                }, 500);
                return true;
            }
            return false;
        """
        result = driver.execute_script(script)
        if result:
            print(f"Clicked on first available card button as fallback")
            time.sleep(3)
            return True
    except Exception as e:
        print(f"Fallback method failed: {e}")
    
    print(f"All methods to click {device_type} card failed")
    return False

def get_dropdown_options(driver, input_id):
    """Retrieve all available options from a dropdown."""
    try:
        dropdown = WebDriverWait(driver, 10).until(
            EC.element_to_be_clickable((By.ID, input_id))
        )
        driver.execute_script("arguments[0].scrollIntoView({block: 'center'});", dropdown)
        dropdown.click()
        time.sleep(1)
        options = driver.find_elements(By.CSS_SELECTOR, f"div[id^='{input_id.replace('input', 'option')}-']")
        option_texts = [option.text.strip() for option in options if option.text.strip()]
        
        # Close dropdown by clicking again
        dropdown.click()
        time.sleep(0.5)
        return option_texts
    except Exception as e:
        print(f"Error getting options for {input_id}: {e}")
        return []

def select_dropdown_option(driver, input_id, option_index, wait, trade_in_data=None):
    """Select an option from a dropdown by index using multiple methods."""
    try:
        print(f"Selecting option {option_index} from dropdown {input_id}")
        
        # Method 1: Standard Selenium approach
        dropdown = wait.until(EC.element_to_be_clickable((By.ID, input_id)))
        driver.execute_script("arguments[0].scrollIntoView({block: 'center'});", dropdown)
        dropdown.click()
        time.sleep(1)
        
        option_selector = f"//div[@id='{input_id.replace('input', 'option')}-{option_index}']"
        option = wait.until(EC.element_to_be_clickable((By.XPATH, option_selector)))
        option_text = option.text.strip()
        
        # Update trade-in data if provided
        if trade_in_data is not None:
            if "react-select-2-input" in input_id:
                trade_in_data["Brand"] = option_text
            elif "react-select-3-input" in input_id:
                trade_in_data["Model"] = option_text
            elif "react-select-4-input" in input_id:
                # For variant information, extract capacity if possible
                variant_text = option_text
                
                # Extract capacity if available (like 128GB or 256GB)
                capacity_match = re.search(r'\d+\s*(?:GB|TB)(?=/|$)|(?<=/)\d+\s*(?:GB|TB)', variant_text)
                if capacity_match:
                    all_matches = re.findall(r'\d+\s*(?:GB|TB)', variant_text)
                    trade_in_data["Capacity"] = all_matches[-1] if all_matches else ""
                    if trade_in_data["Capacity"] == "1024GB":
                        trade_in_data["Capacity"] = "1TB"
                    elif trade_in_data["Capacity"] == "2048GB":
                        trade_in_data["Capacity"] = "2TB"
                else:
                    trade_in_data["Capacity"] = ""
        
        option.click()
        print(f"Selected option: {option_text}")
        return True
    except Exception as e:
        print(f"Standard selection failed: {e}")
        
        # Method 2: JavaScript approach
        try:
            print("Trying JavaScript approach...")
            driver.execute_script(f"""
                var dropdown = document.getElementById('{input_id}');
                if (dropdown) {{
                    dropdown.scrollIntoView({{block: 'center'}});
                    dropdown.click();
                    setTimeout(() => {{
                        var option = document.querySelector('div[id="{input_id.replace("input", "option")}-{option_index}"]');
                        if (option) {{
                            option.click();
                        }}
                    }}, 1000);
                }}
            """)
            time.sleep(2)
            return True
        except Exception as js_error:
            print(f"JavaScript approach failed: {js_error}")
            return False

def handle_popup(driver, wait):
    """Handle the terms & conditions popup."""
    try:
        # Check the checkbox "I hereby agree, understand and wish to proceed"
        checkbox = wait.until(EC.element_to_be_clickable(
            (By.ID, "checked")
        ))
        driver.execute_script("arguments[0].scrollIntoView({block: 'center'});", checkbox)
        checkbox.click()
        print("Clicked the agreement checkbox")
        
        # Click the Proceed button
        proceed_button = wait.until(EC.element_to_be_clickable(
            (By.XPATH, "//button[contains(@class, 'progress-button-next') and text()='Proceed']")
        ))
        driver.execute_script("arguments[0].scrollIntoView({block: 'center'});", proceed_button)
        proceed_button.click()
        print("Clicked the Proceed button")
        time.sleep(2)
        return True
    except Exception as e:
        print(f"Failed to handle T&C popup using standard method: {e}")
        # Fallback method using JavaScript
        try:
            driver.execute_script("""
                var checkbox = document.getElementById('checked');
                if (checkbox) {
                    checkbox.scrollIntoView({block: 'center'});
                    checkbox.click();
                    
                    setTimeout(() => {
                        var proceedBtn = document.querySelector('button.progress-button-next');
                        if (proceedBtn) {
                            proceedBtn.scrollIntoView({block: 'center'});
                            proceedBtn.click();
                        }
                    }, 1000);
                }
            """)
            time.sleep(2)
            return True
        except Exception as js_error:
            print(f"Failed to handle T&C popup using JavaScript: {js_error}")
            return False

def navigate_and_complete_form(driver, wait, device_type, brand_index, model_index, variant_index, screen_condition, output_file):
    """Navigate the website and complete the form for a specific configuration."""
    screen_condition_id = f"LCDS-01-{screen_condition}"
    print(f"Processing: {device_type}, Brand index {brand_index}, Model index {model_index}, Variant index {variant_index}, Condition: {screen_condition}")

    # Initialize data dictionary with the required columns
    trade_in_data = {
        "Country": "Singapore",
        "Device Type": device_type,
        "Brand": "",
        "Model": "",
        "Capacity": "",
        "Color": "",  # Left blank as requested
        "Launch RRP": "",  # Left blank as requested
        "Condition": get_condition_mapping(screen_condition),
        "Value Type": "Trade-in",
        "Currency": "SGD",
        "Value": "",
        "Source": "SG_RV_Source5",
        "Updated on": datetime.now().strftime("%Y-%m-%d"),
        "Updated by": "",  # Left blank as requested
        "Comments": ""  # Left blank as requested
    }

    try:
        driver.get("https://m1tradein.compasia.com/?utm_source=website&utm_medium=cta&utm_campaign=new")
        
        # Click device type button
        click_device_type(driver, wait, device_type)
        time.sleep(2)
        
        # Handle the terms & conditions popup
        handle_popup(driver, wait)

        # Select brand
        select_dropdown_option(driver, "react-select-2-input", brand_index, wait, trade_in_data)
        time.sleep(2)
        
        # Select model
        select_dropdown_option(driver, "react-select-3-input", model_index, wait, trade_in_data)
        time.sleep(2)
        
        # Select variant
        select_dropdown_option(driver, "react-select-4-input", variant_index, wait, trade_in_data)
        time.sleep(2)

        # Click next button
        try:
            next_button = wait.until(EC.element_to_be_clickable(
                (By.XPATH, "//button[contains(@class, 'progress-button-next') and not(@disabled)]")
            ))
            driver.execute_script("arguments[0].scrollIntoView({block: 'center'});", next_button)
            next_button.click()
            print("Successfully clicked Next button")
        except Exception as e:
            print(f"Failed to click Next button: {e}")
            driver.execute_script("""
                var nextBtn = document.querySelector('button.progress-button-next:not([disabled])');
                if (nextBtn) {
                    nextBtn.scrollIntoView({block: 'center'});
                    nextBtn.click();
                }
            """)
            time.sleep(1)

        time.sleep(2)
        
        # NEW CODE: Click Skip button on the new third page
        time.sleep(3)  # Add longer pause to make sure page loads completely
        
        try:
            # Use direct XPATH to the Skip button based on exact text and class
            skip_button = wait.until(EC.element_to_be_clickable(
                (By.XPATH, "//button[contains(@class, 'progress-button-next') and text()='Skip']")
            ))
            driver.execute_script("arguments[0].scrollIntoView({block: 'center'});", skip_button)
            time.sleep(1)  # Add a small pause before clicking
            driver.execute_script("arguments[0].click();", skip_button)  # Click using JavaScript
            print("Successfully clicked Skip button using direct XPATH")
            time.sleep(2)  # Give more time after clicking
        except Exception as e:
            print(f"Failed to click Skip button: {e}")
            # Take a screenshot for debugging if the skip button click fails
            driver.save_screenshot(f"skip_button_error_{device_type}_{brand_index}_{model_index}_{variant_index}.png")

        time.sleep(2)

        # Fill diagnostic form using JavaScript
        driver.execute_script(f"""
            function clickYesButton(labelText) {{
                var labels = document.querySelectorAll('label.diagnostic-form-label');
                for (var i = 0; i < labels.length; i++) {{
                    if (labels[i].textContent.includes(labelText)) {{
                        var yesButton = labels[i].parentNode.querySelector('button:first-of-type');
                        if (yesButton) {{
                            yesButton.scrollIntoView({{block: 'center'}});
                            yesButton.click();
                        }}
                        break;
                    }}
                }}
            }}
            
            // Device locks
            clickYesButton('Is your device free of any locks');
            
            // Screen condition
            var screenLabel = document.querySelector('label[for="{screen_condition_id}"]');
            if (screenLabel) {{
                screenLabel.scrollIntoView({{block: 'center'}});
                screenLabel.click();
            }}
            
            // Body condition
            var bodyLabel = document.querySelector('label[for="DECO-01-flawless"]');
            if (bodyLabel) {{
                bodyLabel.scrollIntoView({{block: 'center'}});
                bodyLabel.click();
            }}
            
            // Other conditions
            clickYesButton('Fingerprint/Face ID working');
            clickYesButton('device functions below working fine');
            clickYesButton('front and back cameras');
            
            // None of the above checkbox
            var labels = document.querySelectorAll('label');
            for (var i = 0; i < labels.length; i++) {{
                if (labels[i].textContent.trim().includes('None of the above')) {{
                    var checkbox = labels[i].previousElementSibling;
                    if (!checkbox || checkbox.type !== 'checkbox') {{
                        var parent = labels[i].parentElement;
                        checkbox = parent.querySelector('input[type="checkbox"]');
                    }}
                    if (checkbox && checkbox.type === 'checkbox') {{
                        checkbox.scrollIntoView({{block: 'center'}});
                        checkbox.checked = true;
                        checkbox.click();
                        checkbox.dispatchEvent(new Event('change', {{ bubbles: true }}));
                    }}
                    break;
                }}
            }}
        """)
        time.sleep(2)

        # Click Get Quote button
        try:
            quote_button = wait.until(EC.element_to_be_clickable(
                (By.XPATH, "//button[@type='submit' and contains(text(), 'Get Quote')]")
            ))
            driver.execute_script("arguments[0].scrollIntoView({block: 'center'});", quote_button)
            time.sleep(0.5)
            quote_button.click()
            print("Clicked Get Quote button")
        except Exception as e:
            print(f"Failed to click Get Quote button: {e}")
            driver.execute_script("""
                var quoteBtn = document.querySelector('button[type="submit"]');
                if (quoteBtn) {
                    quoteBtn.scrollIntoView({block: 'center'});
                    quoteBtn.click();
                }
            """)
            time.sleep(1)

        time.sleep(5)

        # Extract trade-in value
        try:
            wait.until(EC.visibility_of_element_located((By.CLASS_NAME, "pricing-display-table")))
            
            # Currency is set to SGD by default, but extract it if available
            currency_element = driver.find_element(By.CLASS_NAME, "price-product-name.currency")
            if currency_element:
                currency = currency_element.text.strip()
                if currency:
                    trade_in_data["Currency"] = currency

            # Extract the price
            price_element = driver.find_element(By.CLASS_NAME, "pricing-display-price")
            price_text = price_element.text.strip() if price_element else ""
            price_clean = re.sub(r'[^0-9.]', '', price_text)
            trade_in_data["Value"] = price_clean

            print(f"Extracted trade-in value: {trade_in_data['Currency']} {price_clean}")
            save_to_excel(trade_in_data, output_file)
            
            return True

        except Exception as e:
            print(f"Failed to extract trade-in value: {e}")
            driver.save_screenshot(f"error_{device_type}_{brand_index}_{model_index}_{variant_index}_{screen_condition}.png")
            return False

    except Exception as main_error:
        print(f"Main execution failed: {main_error}")
        driver.save_screenshot(f"error_{device_type}_{brand_index}_{model_index}_{variant_index}_{screen_condition}.png")
        return False

def save_to_excel(data, output_file):
    """Save the extracted trade-in data to an Excel file."""
    # Ensure output directory exists
    os.makedirs(os.path.dirname(output_file) if os.path.dirname(output_file) else ".", exist_ok=True)
    
    if os.path.exists(output_file):
        try:
            workbook = openpyxl.load_workbook(output_file)
            sheet = workbook.active
        except:
            workbook = openpyxl.Workbook()
            sheet = workbook.active
            # Use the new column headers
            headers = ["Country", "Device Type", "Brand", "Model", "Capacity", "Color", 
                      "Launch RRP", "Condition", "Value Type", "Currency", "Value", 
                      "Source", "Updated on", "Updated by", "Comments"]
            sheet.append(headers)
    else:
        workbook = openpyxl.Workbook()
        sheet = workbook.active
        # Use the new column headers
        headers = ["Country", "Device Type", "Brand", "Model", "Capacity", "Color", 
                  "Launch RRP", "Condition", "Value Type", "Currency", "Value", 
                  "Source", "Updated on", "Updated by", "Comments"]
        sheet.append(headers)

    # Add new row with all the required columns
    row_data = [
        data["Country"],
        data["Device Type"],
        data["Brand"],
        data["Model"],
        data["Capacity"],
        data["Color"],
        data["Launch RRP"],
        data["Condition"],
        data["Value Type"],
        data["Currency"],
        data["Value"],
        data["Source"],
        data["Updated on"],
        data["Updated by"],
        data["Comments"]
    ]
    sheet.append(row_data)

    try:
        workbook.save(output_file)
        print(f"Saved data to {output_file}")
    except Exception as e:
        print(f"Error saving Excel file: {e}")
        alt_file_name = os.path.join(os.path.dirname(output_file), f"SG_RV_Source5.xlsx")
        try:
            workbook.save(alt_file_name)
            print(f"Saved to {alt_file_name}")
        except Exception as e2:
            print(f"Could not save Excel file: {e2}")

def process_device_type(device_type, n_scrape=None, driver=None, wait=None, output_file=None):
    """Process a specific device type (Smartphone or Tablet)."""
    # brands = ["Apple", "Samsung", "Google", "Huawei", "Xiaomi", "Oppo", "OnePlus", "Sony", "LG", "Motorola", "Vivo", "Realme", "Honor", "Nubia", "Nothing"]
    brands = ["Apple", "Samsung"]
    screen_conditions = ["flawless", "minor_scratches", "cracked"]
    
    # If not provided, set up driver and wait
    if driver is None or wait is None:
        driver = setup_driver()
        ignored_exceptions = (NoSuchElementException, StaleElementReferenceException)
        wait = WebDriverWait(driver, 15, 0.5, ignored_exceptions=ignored_exceptions)
        should_close_driver = True
    else:
        should_close_driver = False

    try:
        # Navigate to website and click on device type
        driver.get("https://m1tradein.compasia.com/?utm_source=website&utm_medium=cta&utm_campaign=new")
        click_device_type(driver, wait, device_type)
        
        # Handle the terms & conditions popup
        handle_popup(driver, wait)

        # Get all brand options and find indices for the brands we're interested in
        brand_options = get_dropdown_options(driver, "react-select-2-input")
        brand_indices = []
        
        for brand in brands:
            try:
                idx = next((i for i, b in enumerate(brand_options) if b == brand), None)
                if idx is not None:
                    brand_indices.append(idx)
                    print(f"Found {brand} at index {idx}")
                else:
                    print(f"Brand {brand} not found in options")
            except Exception as e:
                print(f"Error finding brand {brand}: {e}")
                continue

        # Counter for number of scrapes completed
        scrape_count = 0
        
        # Main loop for all combinations
        for brand_idx in brand_indices:
            # Navigate to the main page for each brand
            driver.get("https://m1tradein.compasia.com/?utm_source=website&utm_medium=cta&utm_campaign=new")
            click_device_type(driver, wait, device_type)
            time.sleep(2)
            
            # Handle the terms & conditions popup
            handle_popup(driver, wait)
            
            # Select the brand
            select_dropdown_option(driver, "react-select-2-input", brand_idx, wait)
            time.sleep(2)
            
            # Get available models for this brand
            model_options = get_dropdown_options(driver, "react-select-3-input")
            num_models = len(model_options)
            print(f"Found {num_models} models for brand index {brand_idx}: {model_options}")
            
            for model_idx in range(num_models):
                # For each model, go back to the main page
                driver.get("https://m1tradein.compasia.com/?utm_source=website&utm_medium=cta&utm_campaign=new")
                click_device_type(driver, wait, device_type)
                time.sleep(2)
                
                # Handle the terms & conditions popup
                handle_popup(driver, wait)
                
                # Select brand again
                select_dropdown_option(driver, "react-select-2-input", brand_idx, wait)
                time.sleep(2)
                
                # Select model
                select_dropdown_option(driver, "react-select-3-input", model_idx, wait)
                time.sleep(2)
                
                # Get variants for this model
                variant_options = get_dropdown_options(driver, "react-select-4-input")
                num_variants = len(variant_options)
                print(f"Found {num_variants} variants for model index {model_idx}: {variant_options}")
                
                for variant_idx in range(num_variants):
                    for condition in screen_conditions:
                        print(f"\nStarting new configuration: {device_type}, Brand idx {brand_idx}, Model idx {model_idx}, Variant idx {variant_idx}, Condition: {condition}")
                        result = navigate_and_complete_form(driver, wait, device_type, brand_idx, model_idx, variant_idx, condition, output_file)
                        time.sleep(2)  # Brief pause between iterations
                        
                        # Increment the scrape counter
                        scrape_count += 1
                        
                        # If there was an error, retry once
                        if not result:
                            print("Retrying after error...")
                            navigate_and_complete_form(driver, wait, device_type, brand_idx, model_idx, variant_idx, condition, output_file)
                            time.sleep(2)
                            
                        # Check if we've reached the requested number of scrapes
                        if n_scrape is not None and scrape_count >= n_scrape:
                            print(f"Completed {scrape_count} scrapes as requested. Stopping.")
                            return scrape_count

        return scrape_count

    except Exception as e:
        print(f"Error in processing {device_type}: {e}")
        driver.save_screenshot(f"{device_type}_loop_error.png")
        return 0
    finally:
        if should_close_driver:
            driver.quit()
            print(f"Browser closed after processing {device_type}.")

def main_loop(n_scrape=None, output_file=None):
    """Main loop to iterate through both device types, brands, models, variants, and screen conditions."""
    # Use default output path if not specified
    if output_file is None:
        # Check for environment variable first
        output_dir = os.environ.get("OUTPUT_DIR", "output")
        os.makedirs(output_dir, exist_ok=True)
        output_file = os.path.join(output_dir, "SG_RV_Source5.xlsx")
    
    print(f"Will save results to: {output_file}")
    
    # Create output directory if it doesn't exist
    os.makedirs(os.path.dirname(output_file) if os.path.dirname(output_file) else ".", exist_ok=True)
    
    # Setup driver (single browser instance for the whole script)
    driver = setup_driver()
    
    ignored_exceptions = (NoSuchElementException, StaleElementReferenceException)
    wait = WebDriverWait(driver, 15, 0.5, ignored_exceptions=ignored_exceptions)

    try:
        # When n_scrape is provided, it applies to each device type individually
        # So we'll scrape n_scrape smartphones AND n_scrape tablets
        
        # Process smartphones
        print("=== Processing Smartphones ===")
        smartphone_scrapes = process_device_type("SmartPhone", n_scrape, driver, wait, output_file)
        print(f"Completed {smartphone_scrapes} smartphone configurations")
        
        # Process tablets
        print("\n=== Processing Tablets ===")
        tablet_scrapes = process_device_type("Tablet", n_scrape, driver, wait, output_file)
        print(f"Completed {tablet_scrapes} tablet configurations")
        
        print(f"Total configurations processed: {smartphone_scrapes + tablet_scrapes}")

    except Exception as e:
        print(f"Error in main loop: {e}")
        driver.save_screenshot("main_loop_error.png")
    finally:
        driver.quit()
        print("Browser closed. Process complete.")

if __name__ == "__main__":
    parser = argparse.ArgumentParser(description='Scrape m1 trade-in values for smartphones and tablets')
    parser.add_argument('-n', type=int, help='Number of scrapes to perform for EACH device type (smartphones and tablets)', default=None)
    parser.add_argument('-o', '--output', type=str, help='Output Excel file path', default=None)
    args = parser.parse_args()
    
    main_loop(args.n, args.output)