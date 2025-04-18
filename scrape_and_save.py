from selenium import webdriver
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.common.exceptions import TimeoutException, NoSuchElementException, StaleElementReferenceException
from webdriver_manager.chrome import ChromeDriverManager
import time
import openpyxl
from datetime import datetime
import os
import re
import subprocess

def setup_driver():
    """Set up the Chrome WebDriver with appropriate options."""
    options = webdriver.ChromeOptions()
    options.add_argument("--start-maximized")
    options.add_argument("--disable-extensions")
    options.add_argument("--disable-popup-blocking")
    options.add_argument("--disable-infobars")
    
    # Add headless mode options when running in GitHub Actions
    if os.environ.get('GITHUB_ACTIONS') == 'true':
        print("Running in GitHub Actions - enabling headless mode")
        options.add_argument('--headless')
        options.add_argument('--no-sandbox')
        options.add_argument('--disable-dev-shm-usage')
        options.add_argument('--disable-gpu')
        options.add_argument('--window-size=1920,1080')
    
    # Use webdriver-manager to handle ChromeDriver
    service = Service(ChromeDriverManager().install())
    driver = webdriver.Chrome(service=service, options=options)
    return driver

def get_dropdown_options(driver, input_id):
    """Retrieve all available options from a dropdown."""
    try:
        dropdown = WebDriverWait(driver, 10).until(
            EC.element_to_be_clickable((By.ID, input_id))
        )
        driver.execute_script("arguments[0].scrollIntoView({block: 'center'});", dropdown)
        dropdown.click()
        time.sleep(1)
        options = driver.find_elements(By.CSS_SELECTOR, f"div[id^='{input_id.replace('input', 'option')}-']")
        option_texts = [option.text.strip() for option in options if option.text.strip()]
        
        # Close dropdown by clicking again
        dropdown.click()
        time.sleep(0.5)
        return option_texts
    except Exception as e:
        print(f"Error getting options for {input_id}: {e}")
        return []

def select_dropdown_option(driver, input_id, option_index, wait, trade_in_data=None):
    """Select an option from a dropdown by index using multiple methods."""
    try:
        print(f"Selecting option {option_index} from dropdown {input_id}")
        
        # Method 1: Standard Selenium approach
        dropdown = wait.until(EC.element_to_be_clickable((By.ID, input_id)))
        driver.execute_script("arguments[0].scrollIntoView({block: 'center'});", dropdown)
        dropdown.click()
        time.sleep(1)
        
        option_selector = f"//div[@id='{input_id.replace('input', 'option')}-{option_index}']"
        option = wait.until(EC.element_to_be_clickable((By.XPATH, option_selector)))
        option_text = option.text.strip()
        
        # Update trade-in data if provided
        if trade_in_data is not None:
            if "react-select-2-input" in input_id:
                trade_in_data["Brand"] = option_text
            elif "react-select-3-input" in input_id:
                trade_in_data["Model"] = option_text
            elif "react-select-4-input" in input_id:
                trade_in_data["Variant"] = option_text
                capacity_match = re.search(r'\d+\s*(?:GB|TB)(?=/|$)|(?<=/)\d+\s*(?:GB|TB)', option_text)
                if capacity_match:
                    all_matches = re.findall(r'\d+\s*(?:GB|TB)', option_text)
                    trade_in_data["Capacity"] = all_matches[-1] if all_matches else "Unknown"
                else:
                    trade_in_data["Capacity"] = "Unknown"
                trade_in_data["Type"] = "Smartphone"
        
        option.click()
        print(f"Selected option: {option_text}")
        return True
    except Exception as e:
        print(f"Standard selection failed: {e}")
        
        # Method 2: JavaScript approach
        try:
            print("Trying JavaScript approach...")
            driver.execute_script(f"""
                var dropdown = document.getElementById('{input_id}');
                if (dropdown) {{
                    dropdown.scrollIntoView({{block: 'center'}});
                    dropdown.click();
                    setTimeout(() => {{
                        var option = document.querySelector('div[id="{input_id.replace("input", "option")}-{option_index}"]');
                        if (option) {{
                            option.click();
                        }}
                    }}, 1000);
                }}
            """)
            time.sleep(2)
            return True
        except Exception as js_error:
            print(f"JavaScript approach failed: {js_error}")
            return False

def commit_changes(brand, model, variant, condition, value):
    """Commit changes to the repository after each row is added."""
    try:
        # Set up git config if running in GitHub Actions
        if os.environ.get('GITHUB_ACTIONS') == 'true':
            subprocess.run(["git", "config", "--global", "user.name", "GitHub Action Bot"])
            subprocess.run(["git", "config", "--global", "user.email", "action@github.com"])
        
        # Add the Excel file
        subprocess.run(["git", "add", "tradein_values.xlsx"])
        
        # Commit with a descriptive message
        commit_message = f"Add trade-in value for {brand} {model} {variant} ({condition}): {value}"
        result = subprocess.run(["git", "commit", "-m", commit_message], 
                                capture_output=True, text=True)
        
        # Check if commit was successful
        if "nothing to commit" in result.stdout or "nothing to commit" in result.stderr:
            print("No changes to commit.")
            return False
        
        # Push the changes immediately
        subprocess.run(["git", "push"])
        print(f"Successfully committed and pushed changes: {commit_message}")
        return True
    
    except Exception as e:
        print(f"Error during git operations: {e}")
        return False

def navigate_and_complete_form(driver, wait, brand_index, model_index, variant_index, screen_condition):
    """Navigate the website and complete the form for a specific configuration."""
    screen_condition_id = f"LCDS-01-{screen_condition}"
    print(f"Processing: Brand index {brand_index}, Model index {model_index}, Variant index {variant_index}, Screen condition: {screen_condition}")

    # Initialize data dictionary
    trade_in_data = {
        "Country": "Singapore",
        "Device": "Smartphone",
        "Type": "Smartphone",
        "Brand": "",
        "Model": "",
        "Variant": "",
        "Capacity": "",
        "Front Condition": screen_condition.replace("_", " ").title(),
        "Value Type": "Trade In",
        "Currency": "",
        "Value": "",
        "Updated on": datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    }

    try:
        driver.get("https://compasiatradeinsg.com/tradein/sell")
        
        # Click Smartphone button
        wait.until(EC.element_to_be_clickable(
            (By.XPATH, "//div[contains(@class, 'card-button-footer') and text()='Smartphone']")
        )).click()
        time.sleep(2)

        # Select brand
        select_dropdown_option(driver, "react-select-2-input", brand_index, wait, trade_in_data)
        time.sleep(2)
        
        # Select model
        select_dropdown_option(driver, "react-select-3-input", model_index, wait, trade_in_data)
        time.sleep(2)
        
        # Select variant
        select_dropdown_option(driver, "react-select-4-input", variant_index, wait, trade_in_data)
        time.sleep(2)

        # Click next button
        try:
            next_button = wait.until(EC.element_to_be_clickable(
                (By.XPATH, "//button[contains(@class, 'progress-button-next') and not(@disabled)]")
            ))
            driver.execute_script("arguments[0].scrollIntoView({block: 'center'});", next_button)
            next_button.click()
            print("Successfully clicked Next button")
        except Exception as e:
            print(f"Failed to click Next button: {e}")
            driver.execute_script("""
                var nextBtn = document.querySelector('button.progress-button-next:not([disabled])');
                if (nextBtn) {
                    nextBtn.scrollIntoView({block: 'center'});
                    nextBtn.click();
                }
            """)
            time.sleep(1)

        time.sleep(2)

        # Fill diagnostic form using JavaScript
        driver.execute_script(f"""
            function clickYesButton(labelText) {{
                var labels = document.querySelectorAll('label.diagnostic-form-label');
                for (var i = 0; i < labels.length; i++) {{
                    if (labels[i].textContent.includes(labelText)) {{
                        var yesButton = labels[i].parentNode.querySelector('button:first-of-type');
                        if (yesButton) {{
                            yesButton.scrollIntoView({{block: 'center'}});
                            yesButton.click();
                        }}
                        break;
                    }}
                }}
            }}
            
            // Device locks
            clickYesButton('Is your device free of any locks');
            
            // Screen condition
            var screenLabel = document.querySelector('label[for="{screen_condition_id}"]');
            if (screenLabel) {{
                screenLabel.scrollIntoView({{block: 'center'}});
                screenLabel.click();
            }}
            
            // Body condition
            var bodyLabel = document.querySelector('label[for="DECO-01-flawless"]');
            if (bodyLabel) {{
                bodyLabel.scrollIntoView({{block: 'center'}});
                bodyLabel.click();
            }}
            
            // Other conditions
            clickYesButton('Fingerprint/Face ID working');
            clickYesButton('device functions below working fine');
            clickYesButton('front and back cameras');
            
            // None of the above checkbox
            var labels = document.querySelectorAll('label');
            for (var i = 0; i < labels.length; i++) {{
                if (labels[i].textContent.trim().includes('None of the above')) {{
                    var checkbox = labels[i].previousElementSibling;
                    if (!checkbox || checkbox.type !== 'checkbox') {{
                        var parent = labels[i].parentElement;
                        checkbox = parent.querySelector('input[type="checkbox"]');
                    }}
                    if (checkbox && checkbox.type === 'checkbox') {{
                        checkbox.scrollIntoView({{block: 'center'}});
                        checkbox.checked = true;
                        checkbox.click();
                        checkbox.dispatchEvent(new Event('change', {{ bubbles: true }}));
                    }}
                    break;
                }}
            }}
        """)
        time.sleep(2)

        # Click Get Quote button
        try:
            quote_button = wait.until(EC.element_to_be_clickable(
                (By.XPATH, "//button[@type='submit' and contains(text(), 'Get Quote')]")
            ))
            driver.execute_script("arguments[0].scrollIntoView({block: 'center'});", quote_button)
            time.sleep(0.5)
            quote_button.click()
            print("Clicked Get Quote button")
        except Exception as e:
            print(f"Failed to click Get Quote button: {e}")
            driver.execute_script("""
                var quoteBtn = document.querySelector('button[type="submit"]');
                if (quoteBtn) {
                    quoteBtn.scrollIntoView({block: 'center'});
                    quoteBtn.click();
                }
            """)
            time.sleep(1)

        time.sleep(5)

        # Extract trade-in value
        try:
            wait.until(EC.visibility_of_element_located((By.CLASS_NAME, "pricing-display-table")))
            
            currency_element = driver.find_element(By.CLASS_NAME, "price-product-name.currency")
            currency = currency_element.text.strip() if currency_element else "SGD"
            trade_in_data["Currency"] = currency

            price_element = driver.find_element(By.CLASS_NAME, "pricing-display-price")
            price_text = price_element.text.strip() if price_element else ""
            price_clean = re.sub(r'[^0-9.]', '', price_text)
            trade_in_data["Value"] = price_clean

            print(f"Extracted trade-in value: {currency} {price_clean}")
            save_to_excel(trade_in_data)
            
            # Commit changes to the repository after each successful scrape
            commit_changes(
                trade_in_data["Brand"], 
                trade_in_data["Model"], 
                trade_in_data["Variant"], 
                trade_in_data["Front Condition"], 
                f"{trade_in_data['Currency']} {trade_in_data['Value']}"
            )
            
            return True

        except Exception as e:
            print(f"Failed to extract trade-in value: {e}")
            driver.save_screenshot(f"error_{brand_index}_{model_index}_{variant_index}_{screen_condition}.png")
            return False

    except Exception as main_error:
        print(f"Main execution failed: {main_error}")
        driver.save_screenshot(f"error_{brand_index}_{model_index}_{variant_index}_{screen_condition}.png")
        return False

def save_to_excel(data):
    """Save the extracted trade-in data to an Excel file."""
    file_name = "tradein_values.xlsx"
    if os.path.exists(file_name):
        try:
            workbook = openpyxl.load_workbook(file_name)
            sheet = workbook.active
        except:
            workbook = openpyxl.Workbook()
            sheet = workbook.active
            headers = ["Country", "Device Type", "Brand", "Model", "Variant", "Capacity",
                       "Front Condition", "Value Type", "Currency", "Value", "Updated on"]
            sheet.append(headers)
    else:
        workbook = openpyxl.Workbook()
        sheet = workbook.active
        headers = ["Country", "Device Type", "Brand", "Model", "Variant", "Capacity",
                   "Front Condition", "Value Type", "Currency", "Value", "Updated on"]
        sheet.append(headers)

    row_data = [
        data["Country"],
        "Smartphone",
        data["Brand"],
        data["Model"],
        data.get("Variant", ""),
        data["Capacity"],
        data["Front Condition"],
        "Trade In",
        data["Currency"],
        data["Value"],
        data["Updated on"]
    ]
    sheet.append(row_data)

    try:
        workbook.save(file_name)
        print(f"Saved data to {file_name}")
    except Exception as e:
        print(f"Error saving Excel file: {e}")
        alt_file_name = f"tradein_values_{int(time.time())}.xlsx"
        try:
            workbook.save(alt_file_name)
            print(f"Saved to {alt_file_name}")
        except:
            print("Could not save Excel file")

def main_loop():
    """Main loop to iterate through brands, models, variants, and screen conditions."""
    brands = ["Apple", "Google"]
    screen_conditions = ["flawless", "minor_scratches", "cracked"]

    # Setup driver (single browser instance)
    driver = setup_driver()
    
    ignored_exceptions = (NoSuchElementException, StaleElementReferenceException)
    wait = WebDriverWait(driver, 15, 0.5, ignored_exceptions=ignored_exceptions)

    try:
        # Take screenshot of initial state for debugging
        if os.environ.get('GITHUB_ACTIONS') == 'true':
            driver.save_screenshot("initial_page.png")
            
        # Navigate to website and click on Smartphone
        driver.get("https://compasiatradeinsg.com/tradein/sell")
        wait.until(EC.element_to_be_clickable(
            (By.XPATH, "//div[contains(@class, 'card-button-footer') and text()='Smartphone']")
        )).click()
        time.sleep(2)

        # Get all brand options and find indices for the brands we're interested in
        brand_options = get_dropdown_options(driver, "react-select-2-input")
        brand_indices = []
        
        for brand in brands:
            try:
                idx = next((i for i, b in enumerate(brand_options) if b == brand), None)
                if idx is not None:
                    brand_indices.append(idx)
                    print(f"Found {brand} at index {idx}")
                else:
                    print(f"Brand {brand} not found in options")
            except Exception as e:
                print(f"Error finding brand {brand}: {e}")
                continue

        # Main loop for all combinations
        for brand_idx in brand_indices:
            # Navigate to the main page for each brand
            driver.get("https://compasiatradeinsg.com/tradein/sell")
            wait.until(EC.element_to_be_clickable(
                (By.XPATH, "//div[contains(@class, 'card-button-footer') and text()='Smartphone']")
            )).click()
            time.sleep(2)
            
            # Select the brand
            select_dropdown_option(driver, "react-select-2-input", brand_idx, wait)
            time.sleep(2)
            
            # Get available models for this brand
            model_options = get_dropdown_options(driver, "react-select-3-input")
            num_models = len(model_options)
            print(f"Found {num_models} models for brand index {brand_idx}: {model_options}")
            
            for model_idx in range(num_models):
                # For each model, go back to the main page
                driver.get("https://compasiatradeinsg.com/tradein/sell")
                wait.until(EC.element_to_be_clickable(
                    (By.XPATH, "//div[contains(@class, 'card-button-footer') and text()='Smartphone']")
                )).click()
                time.sleep(2)
                
                # Select brand again
                select_dropdown_option(driver, "react-select-2-input", brand_idx, wait)
                time.sleep(2)
                
                # Select model
                select_dropdown_option(driver, "react-select-3-input", model_idx, wait)
                time.sleep(2)
                
                # Get variants for this model
                variant_options = get_dropdown_options(driver, "react-select-4-input")
                num_variants = len(variant_options)
                print(f"Found {num_variants} variants for model index {model_idx}: {variant_options}")
                
                for variant_idx in range(num_variants):
                    for condition in screen_conditions:
                        print(f"\nStarting new configuration: Brand idx {brand_idx}, Model idx {model_idx}, Variant idx {variant_idx}, Condition: {condition}")
                        result = navigate_and_complete_form(driver, wait, brand_idx, model_idx, variant_idx, condition)
                        time.sleep(2)  # Brief pause between iterations
                        
                        # If there was an error, retry once
                        if not result:
                            print("Retrying after error...")
                            navigate_and_complete_form(driver, wait, brand_idx, model_idx, variant_idx, condition)
                            time.sleep(2)

    except Exception as e:
        print(f"Error in main loop: {e}")
        driver.save_screenshot("main_loop_error.png")
    finally:
        driver.quit()
        print("Browser closed. Process complete.")

if __name__ == "__main__":
    main_loop()